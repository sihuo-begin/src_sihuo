
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
class ReportBuilder:
    def __init__(self,tpl,out):
        self.wb=load_workbook(tpl); self.ws=self.wb.active; self.out=out
    def insert_images(self,data,dir):
        r=2
        for item in data:
            p=os.path.join(dir,f"{item}.png")
            if os.path.exists(p):
                img=Image(p); img.width=350; img.height=250
                self.ws.add_image(img,f"G{r}")
            p2=os.path.join(dir,f"{item}_GRR.png")
            if os.path.exists(p2):
                img2=Image(p2); img2.width=350; img2.height=250
                self.ws.add_image(img2,f"J{r}")
            self.ws[f"A{r}"]=item
            r+=1
        self.wb.save(self.out)
