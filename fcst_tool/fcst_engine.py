"""
FCST Waterfall Engine - 可复用模块
========================================
封装 FCST 数据解析 + 比对报表生成的全部逻辑。

核心API:
  parse_fcst_xlsx(wk21_path, wk22_path, wk23_path) -> dict
      解析 3 份 xlsx FCST 报表, 返回 {snap_label: [sku_dict, ...]}

  build_report(all_data, weeks, output_path, progress_cb=None) -> None
      生成完整比对 Excel 报表(Summary + 7 个分类 Sheet + SKU)

数据流:
  3 xlsx (用户上传) -> parse_fcst_xlsx -> all_data dict
                       |
                       v
                  build_report -> .xlsx
"""
import re
import pickle
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

SNAPSHOTS = [
    ('WK21', 'SF04 WK21'),
    ('WK22', 'SF05 WK22'),
    ('WK23', 'SF05 WK23'),
]


# ============== 解析部分 ==============
def parse_fcst_xlsx(wk21_path, wk22_path, wk23_path, progress_cb=None):
    """
    解析 3 份 xlsx FCST 报表, 返回 all_data dict
    all_data[snap_label] = [ {version, item_group, market, sku, core_color, monthly:[12], total}, ... ]
    """
    files = {
        'WK21': wk21_path,
        'WK22': wk22_path,
        'WK23': wk23_path,
    }
    all_data = {}
    for label, fpath in files.items():
        if progress_cb:
            progress_cb(f'正在解析 {label} ...')
        data, weeks = _parse_one(fpath, progress_cb)
        all_data[label] = data
        if progress_cb:
            progress_cb(f'  {label}: {len(data)} 行, Total={sum(d["total"] for d in data):,.0f}')
    return all_data, weeks


def _parse_one(fpath, progress_cb=None):
    """解析单个 xlsx 文件"""
    wb = load_workbook(fpath, data_only=True, read_only=True, keep_links=False)
    ws = wb['2026']
    rows = list(ws.iter_rows(values_only=True))
    if progress_cb:
        progress_cb(f'    读取 {len(rows)} 行')

    # row2/row4: 周→月映射
    row2 = rows[1] if len(rows) > 1 else []
    row4 = rows[3] if len(rows) > 3 else []
    week_to_month = {}
    week_labels = []
    for col_idx, week_label in enumerate(row2):
        if col_idx < 25 or not week_label:
            continue
        month_label = row4[col_idx] if col_idx < len(row4) else None
        if not month_label or not isinstance(month_label, str):
            continue
        for m_idx, m_name in enumerate(MONTH_NAMES):
            if m_name in month_label:
                week_to_month[col_idx] = m_idx
                if str(week_label) not in week_labels:
                    week_labels.append(str(week_label))
                break
    if progress_cb:
        progress_cb(f'    周→月映射: {len(week_to_month)} 个周, 共 {len(week_labels)} 个周标签')

    # row5+: 数据行
    sku_data = []
    for row in rows[5:]:
        if not row or not row[0] or not row[8]:
            continue
        sku_data.append({
            'version': str(row[2]) if len(row) > 2 and row[2] else '',
            'item_group': str(row[10]) if len(row) > 10 and row[10] else '',
            'market': str(row[4]) if len(row) > 4 and row[4] else '',
            'sku': str(row[8]) if len(row) > 8 and row[8] else '',
            'core_color': str(row[13]) if len(row) > 13 and row[13] else '',
            'row': row,
        })

    # 月度小计 + total
    for item in sku_data:
        row = item['row']
        monthly = [0.0] * 12
        total = 0.0
        for col_idx in range(25, len(row)):
            val = row[col_idx]
            if val and isinstance(val, (int, float)):
                total += val
                m_idx = week_to_month.get(col_idx)
                if m_idx is not None and 0 <= m_idx < 12:
                    monthly[m_idx] += val
        item['monthly'] = monthly
        item['total'] = total
        del item['row']
    wb.close()
    return sku_data, week_labels


# ============== 工具函数 ==============
def week_to_date(week_str):
    m = re.match(r'(\d{4})-W(\d+)', str(week_str))
    if not m:
        return str(week_str)
    year, wn = int(m.group(1)), int(m.group(2))
    try:
        d = datetime.strptime(f'{year}-W{wn:02d}-1', '%Y-W%W-%w')
        return d.strftime('%Y-%m-%d')
    except Exception:
        return str(week_str)


def get_month_year(week_str):
    m = re.match(r'(\d{4})-W(\d+)', str(week_str))
    if not m:
        return ''
    year, wn = int(m.group(1)), int(m.group(2))
    month_idx = min((wn - 1) * 12 // 52, 11)
    return f"{MONTH_NAMES[month_idx]}'{str(year)[-2:]}"


def month_labels_dedup(weeks):
    seen = set()
    out = []
    for w in weeks:
        ml = get_month_year(w)
        if ml not in seen:
            seen.add(ml)
            out.append(ml)
    return out[:18]


# ============== 样式 ==============
def _styles():
    s = {}
    s['title_font'] = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
    s['title_fill'] = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    s['section_font'] = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    s['section_blue_fill'] = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')
    s['header_font'] = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    s['header_fill'] = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    s['subheader_font'] = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    s['subheader_fill'] = PatternFill(start_color='8AB4D8', end_color='8AB4D8', fill_type='solid')
    s['data_font'] = Font(name='Calibri', size=9, color='000000')
    s['snap_font'] = Font(name='Calibri', size=9, bold=True, color='000000')
    s['snap_fill_wk21'] = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    s['snap_fill_wk22'] = PatternFill(start_color='EAF5EA', end_color='EAF5EA', fill_type='solid')
    s['snap_fill_wk23'] = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
    s['delta_fill'] = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    s['delta_font'] = Font(name='Calibri', size=9, italic=True, color='606060')
    s['total_fill'] = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    s['pos_fill'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    s['neg_fill'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    s['thin'] = Side(style='thin', color='D0D0D0')
    s['border'] = Border(left=s['thin'], right=s['thin'], top=s['thin'], bottom=s['thin'])
    s['center'] = Alignment(horizontal='center', vertical='center', wrap_text=True)
    s['left_align'] = Alignment(horizontal='left', vertical='center', wrap_text=True)
    s['right_align'] = Alignment(horizontal='right', vertical='center')
    s['SNAP_FILLS'] = {'WK21': s['snap_fill_wk21'], 'WK22': s['snap_fill_wk22'], 'WK23': s['snap_fill_wk23']}
    return s


# ============== 分组聚合 ==============
def build_groups(snap_data, group_dims, allowed_ig):
    groups = defaultdict(lambda: {'weekly': defaultdict(float), 'monthly': defaultdict(float), 'total': 0.0, 'weeks': defaultdict(float)})
    for d in snap_data:
        if allowed_ig is not None and d['item_group'] not in allowed_ig:
            continue
        key = tuple(str(d.get(dim, '')) for dim in group_dims)
        for m_idx, m_val in enumerate(d['monthly']):
            if m_val:
                ml = f"{MONTH_NAMES[m_idx]}'26"
                groups[key]['monthly'][ml] += m_val
                groups[key]['total'] += m_val
    return groups


# ============== 报表写入 ==============
def build_report(all_data, weeks, output_path, progress_cb=None):
    """
    生成完整比对 Excel 报表
    all_data: {snap_label: [sku_dict, ...]}  (来自 parse_fcst_xlsx)
    weeks: 周标签列表
    output_path: 生成的 xlsx 路径
    """
    S = _styles()
    unique_months = month_labels_dedup(weeks)
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Summary
    if progress_cb: progress_cb('生成 Summary Sheet...')
    _build_summary(wb, all_data, S)

    # 2-7. 分类 Sheet
    classifications = [
        ('V_Kit_Holder', ['version'], ['VERSION CODE', 'ITEM GROUP'], ['Kit', 'Holder'], 'Holder+Kit'),
        ('V_Kit_PC', ['version'], ['VERSION CODE', 'ITEM GROUP'], ['Kit', 'Pocket Charger'], 'Kit+Pocket Charger'),
        ('V_Kit_Holder_Color', ['version', 'core_color'], ['VERSION CODE', 'ITEM GROUP', 'COLOR'], ['Kit', 'Holder'], 'Holder+Kit'),
        ('V_Kit_PC_Color', ['version', 'core_color'], ['VERSION CODE', 'ITEM GROUP', 'COLOR'], ['Kit', 'Pocket Charger'], 'Kit+Pocket Charger'),
        ('V_Kit_Holder_Market', ['version', 'market'], ['VERSION CODE', 'ITEM GROUP', 'MARKET'], ['Kit', 'Holder'], 'Holder+Kit'),
        ('V_Kit_PC_Market', ['version', 'market'], ['VERSION CODE', 'ITEM GROUP', 'MARKET'], ['Kit', 'Pocket Charger'], 'Kit+Pocket Charger'),
    ]
    for sn, dims, labels, ig, combined in classifications:
        if progress_cb: progress_cb(f'生成 {sn} ...')
        _write_classification_sheet(wb, sn, dims, labels, ig, combined, all_data, weeks, unique_months, S)

    # 8. SKU 全量
    if progress_cb: progress_cb('生成 SKU Sheet ...')
    _write_classification_sheet(
        wb, 'SKU',
        ['sku', 'version', 'item_group', 'core_color', 'market'],
        ['SKU', 'VERSION CODE', 'ITEM GROUP', 'COLOR', 'MARKET'],
        None, None,
        all_data, weeks, unique_months, S
    )

    wb.save(output_path)
    if progress_cb: progress_cb(f'✅ 已保存: {output_path}')


def _write_classification_sheet(wb, sheet_name, group_dims, group_labels, allowed_ig, combined_ig_label, all_data, weeks, unique_months, S):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    snap_groups = {sl: build_groups(all_data[sl], group_dims, allowed_ig) for sl, _ in SNAPSHOTS}
    all_keys = set()
    for g in snap_groups.values():
        all_keys.update(g.keys())
    sorted_keys = sorted(all_keys)

    n_dims = len(group_labels)
    col_total_qty = n_dims + 1
    col_fcst_ver = n_dims + 2
    col_data_start = n_dims + 3

    # 列宽
    widths = [14, 16, 18, 14, 14] + [11] * max(len(weeks), len(unique_months))
    for i, w in enumerate(widths[:col_data_start + max(len(weeks), len(unique_months))]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = f'{get_column_letter(col_data_start)}4'

    # ============== 按周比较表 ==============
    row = 1
    last_col = col_data_start + len(weeks) - 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=f'【按周比较】 {sheet_name}')
    c.font = S['title_font']
    c.fill = S['title_fill']
    c.alignment = S['center']
    ws.row_dimensions[row].height = 28
    row += 2

    # 表头
    headers = group_labels + ['TOTAL QTY', 'FCST Version']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = S['header_font']
        c.fill = S['header_fill']
        c.alignment = S['center']
        c.border = S['border']
    for w_idx, w in enumerate(weeks):
        c = ws.cell(row=row, column=col_data_start + w_idx, value=week_to_date(w))
        c.font = S['subheader_font']
        c.fill = S['subheader_fill']
        c.alignment = S['center']
        c.border = S['border']
    row += 1

    # 数据
    for key in sorted_keys:
        snap_data_rows = []
        for snap_idx, (sl, sl_label) in enumerate(SNAPSHOTS):
            g = snap_groups[sl].get(key, {'total': 0, 'monthly': {}})
            row_vals = [g['total']]
            for w in weeks:
                v = 0.0
                for m_idx, m_name in enumerate(MONTH_NAMES):
                    ml = f"{m_name}'26"
                    if ml == get_month_year(w):
                        v = g['monthly'].get(ml, 0.0)
                        break
                row_vals.append(v)
            snap_data_rows.append(row_vals)

        # 5行: WK21/22/23/Δ12/Δ23
        for offset in range(5):
            dim_i = 0
            for i, lbl in enumerate(group_labels):
                if combined_ig_label is not None and lbl == 'ITEM GROUP':
                    c = ws.cell(row=row, column=1 + i, value=combined_ig_label)
                else:
                    c = ws.cell(row=row, column=1 + i, value=key[dim_i])
                    dim_i += 1
                c.alignment = S['left_align']
                c.border = S['border']
                if offset < 3:
                    sl = SNAPSHOTS[offset][0]
                    c.font = S['data_font']
                    c.fill = S['SNAP_FILLS'][sl]
                else:
                    c.font = S['delta_font']
                    c.fill = S['delta_fill']

            # TOTAL QTY
            if offset < 3:
                tval = snap_data_rows[offset][0]
            elif offset == 3:
                tval = snap_data_rows[1][0] - snap_data_rows[0][0]
            else:
                tval = snap_data_rows[2][0] - snap_data_rows[1][0]
            c = ws.cell(row=row, column=col_total_qty, value=tval)
            c.alignment = S['right_align']
            c.border = S['border']
            c.number_format = '#,##0'
            if offset < 3:
                sl = SNAPSHOTS[offset][0]
                c.font = S['snap_font']
                c.fill = S['SNAP_FILLS'][sl]
            else:
                c.font = S['delta_font']
                c.fill = S['delta_fill']
                if tval > 0:
                    c.font = Font(name='Calibri', size=9, italic=True, color='006100')
                elif tval < 0:
                    c.font = Font(name='Calibri', size=9, italic=True, color='9C0006')

            # FCST Version
            if offset < 3:
                fcst_label = SNAPSHOTS[offset][1]
            else:
                fcst_label = f'Δ {SNAPSHOTS[offset-3][0][2:]}→{SNAPSHOTS[offset-2][0][2:]}'
            c = ws.cell(row=row, column=col_fcst_ver, value=fcst_label)
            c.alignment = S['center']
            c.border = S['border']
            if offset < 3:
                sl = SNAPSHOTS[offset][0]
                c.font = S['snap_font']
                c.fill = S['SNAP_FILLS'][sl]
            else:
                c.font = S['delta_font']
                c.fill = S['delta_fill']

            # 周次数据
            for w_idx, w in enumerate(weeks):
                col = col_data_start + w_idx
                if offset < 3:
                    val = snap_data_rows[offset][1 + w_idx]
                elif offset == 3:
                    val = snap_data_rows[1][1 + w_idx] - snap_data_rows[0][1 + w_idx]
                else:
                    val = snap_data_rows[2][1 + w_idx] - snap_data_rows[1][1 + w_idx]
                c = ws.cell(row=row, column=col, value=val if val else None)
                c.alignment = S['right_align']
                c.border = S['border']
                c.number_format = '#,##0;-#,##0;-'
                if offset < 3:
                    sl = SNAPSHOTS[offset][0]
                    c.font = S['data_font']
                    c.fill = S['SNAP_FILLS'][sl]
                else:
                    c.font = S['delta_font']
                    c.fill = S['delta_fill']
                    if val > 0:
                        c.font = Font(name='Calibri', size=9, italic=True, color='006100')
                    elif val < 0:
                        c.font = Font(name='Calibri', size=9, italic=True, color='9C0006')
            row += 1

    # ============== 按月比较表 ==============
    row += 2
    last_col_m = col_data_start + len(unique_months) - 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col_m)
    c = ws.cell(row=row, column=1, value=f'【按月比较】 {sheet_name}')
    c.font = S['title_font']
    c.fill = S['title_fill']
    c.alignment = S['center']
    ws.row_dimensions[row].height = 28
    row += 2

    headers = group_labels + ['TOTAL QTY', 'FCST Version']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = S['header_font']
        c.fill = S['header_fill']
        c.alignment = S['center']
        c.border = S['border']
    for m_idx, ml in enumerate(unique_months):
        c = ws.cell(row=row, column=col_data_start + m_idx, value=ml)
        c.font = S['subheader_font']
        c.fill = S['subheader_fill']
        c.alignment = S['center']
        c.border = S['border']
    row += 1

    for key in sorted_keys:
        snap_data_rows = []
        for snap_idx, (sl, sl_label) in enumerate(SNAPSHOTS):
            g = snap_groups[sl].get(key, {'total': 0, 'monthly': {}})
            row_vals = [g['total']]
            for ml in unique_months:
                row_vals.append(g['monthly'].get(ml, 0.0))
            snap_data_rows.append(row_vals)

        for offset in range(5):
            dim_i = 0
            for i, lbl in enumerate(group_labels):
                if combined_ig_label is not None and lbl == 'ITEM GROUP':
                    c = ws.cell(row=row, column=1 + i, value=combined_ig_label)
                else:
                    c = ws.cell(row=row, column=1 + i, value=key[dim_i])
                    dim_i += 1
                c.alignment = S['left_align']
                c.border = S['border']
                if offset < 3:
                    sl = SNAPSHOTS[offset][0]
                    c.font = S['data_font']
                    c.fill = S['SNAP_FILLS'][sl]
                else:
                    c.font = S['delta_font']
                    c.fill = S['delta_fill']

            if offset < 3:
                tval = snap_data_rows[offset][0]
            elif offset == 3:
                tval = snap_data_rows[1][0] - snap_data_rows[0][0]
            else:
                tval = snap_data_rows[2][0] - snap_data_rows[1][0]
            c = ws.cell(row=row, column=col_total_qty, value=tval)
            c.alignment = S['right_align']
            c.border = S['border']
            c.number_format = '#,##0'
            if offset < 3:
                sl = SNAPSHOTS[offset][0]
                c.font = S['snap_font']
                c.fill = S['SNAP_FILLS'][sl]
            else:
                c.font = S['delta_font']
                c.fill = S['delta_fill']
                if tval > 0:
                    c.font = Font(name='Calibri', size=9, italic=True, color='006100')
                elif tval < 0:
                    c.font = Font(name='Calibri', size=9, italic=True, color='9C0006')

            if offset < 3:
                fcst_label = SNAPSHOTS[offset][1]
            else:
                fcst_label = f'Δ {SNAPSHOTS[offset-3][0][2:]}→{SNAPSHOTS[offset-2][0][2:]}'
            c = ws.cell(row=row, column=col_fcst_ver, value=fcst_label)
            c.alignment = S['center']
            c.border = S['border']
            if offset < 3:
                sl = SNAPSHOTS[offset][0]
                c.font = S['snap_font']
                c.fill = S['SNAP_FILLS'][sl]
            else:
                c.font = S['delta_font']
                c.fill = S['delta_fill']

            for m_idx, ml in enumerate(unique_months):
                col = col_data_start + m_idx
                if offset < 3:
                    val = snap_data_rows[offset][1 + m_idx]
                elif offset == 3:
                    val = snap_data_rows[1][1 + m_idx] - snap_data_rows[0][1 + m_idx]
                else:
                    val = snap_data_rows[2][1 + m_idx] - snap_data_rows[1][1 + m_idx]
                c = ws.cell(row=row, column=col, value=val if val else None)
                c.alignment = S['right_align']
                c.border = S['border']
                c.number_format = '#,##0;-#,##0;-'
                if offset < 3:
                    sl = SNAPSHOTS[offset][0]
                    c.font = S['data_font']
                    c.fill = S['SNAP_FILLS'][sl]
                else:
                    c.font = S['delta_font']
                    c.fill = S['delta_fill']
                    if val > 0:
                        c.font = Font(name='Calibri', size=9, italic=True, color='006100')
                    elif val < 0:
                        c.font = Font(name='Calibri', size=9, italic=True, color='9C0006')
            row += 1


def _build_summary(wb, all_data, S):
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 24
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 16

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value='FCST 比对 - 项目工程师标准格式')
    c.font = S['title_font']
    c.fill = S['title_fill']
    c.alignment = S['center']
    ws.row_dimensions[row].height = 32
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value='基于 3 个 Snapshot (WK21 SF04 / WK22 SF05 / WK23 SF05) 横向对比').font = Font(italic=True, color='606060')
    row += 2

    # 整体对比
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value='① 整体 FCST 比对')
    c.font = S['section_font']
    c.fill = S['section_blue_fill']
    c.alignment = S['left_align']
    ws.row_dimensions[row].height = 25
    row += 1

    headers = ['指标', 'WK21 SF04', 'WK22 SF05', 'WK23 SF05', 'ΔWK21→22', 'Δ%', 'ΔWK22→23', 'Δ%']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = S['header_font']
        c.fill = S['header_fill']
        c.alignment = S['center']
        c.border = S['border']
    row += 1

    ig_groups = ['Kit', 'Holder', 'Pocket Charger', 'USB Cable', 'Other', 'TOTAL']
    for ig in ig_groups:
        vals = []
        for sl, _ in SNAPSHOTS:
            if ig == 'TOTAL':
                v = sum(d['total'] for d in all_data[sl])
            else:
                v = sum(d['total'] for d in all_data[sl] if d['item_group'] == ig)
            vals.append(v)
        d12 = vals[1] - vals[0]
        d23 = vals[2] - vals[1]
        p12 = d12 / vals[0] if vals[0] else 0
        p23 = d23 / vals[1] if vals[1] else 0
        row_data = [ig, vals[0], vals[1], vals[2], d12, p12, d23, p23]
        for i, v in enumerate(row_data):
            c = ws.cell(row=row, column=i + 1, value=v)
            c.border = S['border']
            c.alignment = S['center'] if i in (0, 5, 7) else S['right_align']
            if i in (5, 7):
                c.number_format = '0.00%'
            elif i > 0:
                c.number_format = '#,##0'
            if ig == 'TOTAL':
                c.fill = S['total_fill']
                c.font = Font(bold=True)
        row += 1

    row += 1
    # Sheet 索引
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value='② Sheet 索引')
    c.font = S['section_font']
    c.fill = S['section_blue_fill']
    c.alignment = S['left_align']
    ws.row_dimensions[row].height = 25
    row += 1

    descs = [
        ('V_Kit_Holder', 'VERSION + ITEM GROUP (Kit/Holder)'),
        ('V_Kit_PC', 'VERSION + ITEM GROUP (Kit/PC)'),
        ('V_Kit_Holder_Color', 'VERSION + ITEM GROUP + COLOR (Kit/Holder)'),
        ('V_Kit_PC_Color', 'VERSION + ITEM GROUP + COLOR (Kit/PC)'),
        ('V_Kit_Holder_Market', 'VERSION + ITEM GROUP + MARKET (Kit/Holder)'),
        ('V_Kit_PC_Market', 'VERSION + ITEM GROUP + MARKET (Kit/PC)'),
        ('SKU', 'SKU + VERSION + ITEM GROUP + COLOR + MARKET (按SKU全量明细)'),
    ]
    for i, (sn, dim) in enumerate(descs):
        ws.cell(row=row, column=1, value=f'{i+1}.').alignment = S['center']
        c2 = ws.cell(row=row, column=2, value=sn)
        c2.font = Font(name='Calibri', size=10, bold=True, color='0066CC')
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row=row, column=3, value=dim)
        row += 1


# ============== CLI 入口 (兼容旧调用) ==============
def main_cli():
    import sys
    if len(sys.argv) < 4:
        print('Usage: fcst_engine.py <wk21.xlsx> <wk22.xlsx> <wk23.xlsx> [output.xlsx]')
        sys.exit(1)
    wk21, wk22, wk23 = sys.argv[1], sys.argv[2], sys.argv[3]
    output = sys.argv[4] if len(sys.argv) > 4 else '/home/workspace/FCST_Waterfall.xlsx'

    def log(msg):
        print(msg)

    all_data, weeks = parse_fcst_xlsx(wk21, wk22, wk23, log)
    build_report(all_data, weeks, output, log)
    print(f'\n✅ 报表已生成: {output}')


if __name__ == '__main__':
    main_cli()
