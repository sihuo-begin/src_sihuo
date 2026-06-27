# ──────────────────────────────────────────────
#  JSON Log Parser
#  Parses MT7 GRR Jason logs → intermediate HVTE-M600099_GRR_data.xlsx
# ──────────────────────────────────────────────
import json
import logging
import glob
import math
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from utils.config import OUTPUT_DIR

import pandas as pd
import os
import json

logger = logging.getLogger(__name__)

# LED column mapping: JSON key → standard column name
LED_JSON_KEYS = {
    "LED_RED_D303":     "LED_RED_D303_INTENSITY",
    "LED_RED_D304":     "LED_RED_D304_INTENSITY",
    "LED_WHITE_D305":   "LED_WHITE_D305_INTENSITY",
    "LED_WHITE_D306":   "LED_WHITE_D306_INTENSITY",
    "LED_WHITE_D307":   "LED_WHITE_D307_INTENSITY",
    "LED_WHITE_D308":   "LED_WHITE_D308_INTENSITY",
    "LED_WHITE_D309":   "LED_WHITE_D309_INTENSITY",
    "LED_AMBER_D310":   "LED_AMBER_D310_INTENSITY",
    "LED_WHITE_D311":   "LED_WHITE_D311_INTENSITY",
}

# All columns in the intermediate GRR Excel
GRR_COLUMNS = [
    "sn",
    "status",
    "failitem",
    "errorcode",
    "starttime",
    "endtime",
    "appraiser",       # operator ID, extracted from QR_SCAN
    "DETECT_DUT",
    "INITIALIZE_CONFIGS",
    "QR_SCAN",
    "TBB_UID",
    "READ_DUSN",
    "SOFTWARE_REVISION",
    "READ_MODEL_NUMBER",
    "DGS_BATTERY_DATA",
    "READ_MT_MODE",
    "APPLICATION_ID_(MT)",
    "RETRIEVE_SBL_MAJOR_ID",
    "HOLDER_PRESENCE",
    "USB_C_INFO",
    "USB_ORIENTATION_A",
    "USB_CC_ADC_SIDE_A_TEST",
    "USB_ORIENTATION_B",
    "USB_CC_ADC_SIDE_B_TEST",
    "BIST_TEST",
    "KEY_LENGTH",
    "KEY_DATA",
    "LED_RED_D303",
    "LED_RED_D304",
    "LED_WHITE_D305",
    "LED_WHITE_D306",
    "LED_WHITE_D307",
    "LED_WHITE_D308",
    "LED_WHITE_D309",
    "LED_AMBER_D310",
    "LED_WHITE_D311",
    "BLE_DISCOVERY",
    "BLE_MAC_ADDRESS_RETRIEVE",
    "RETRIEVE_BLE_STACK_VERSION",
    "SET_BATTERY_CONFIGURATION",
    "BUTTER_RELEASE_TEST",
    "BUTTER_PRESS_TEST",
    "HARD_RESET_TEST",
    "CODENTIFY_CODE_INTO_DUT",
    "BATTERY_SN_INTO_DUT",
    "BATTERY_SN_IN_DUT",
    "RESET_DUT",
    "UTC_TIME_SYNC",
    "APPLICATION_ID_(EXIT_MT)",
    "CODENTIFY_CODE_IN_DUT",
    "READ_DEVICE_BATTERY_CONFIG",
    "DEVICE_SYSTEM_ERROR",
    "BLE_STATUS",
    "OVERALL_TEST_RESULT",
]


def _deep_get(obj: dict, *keys, default=None):
    """Safely navigate nested dict."""
    for k in keys:
        if isinstance(obj, dict):
            obj = obj.get(k, default)
        else:
            return default
    return obj


def _deep_get(obj, *keys, default=None):
    """Safely navigate nested dicts."""
    for k in keys:
        if not isinstance(obj, dict):
            return default
        obj = obj.get(k, default)
        if obj is None:
            return default
    return obj


def _extract_appraiser(test_array: list) -> str:
    """Extract operator/appraiser ID from QR_SCAN entry in test array."""
    for entry in test_array:
        ref = str(_deep_get(entry, "reference") or "").strip()
        if ref == "QR_SCAN":
            val = str(_deep_get(entry, "actual") or "").strip()
            if val:
                return val
    return "UNKNOWN"


def _extract_sn(data: dict) -> str:
    """Extract device serial number from %device_id.device field.
    Prepend ' to prevent Excel from converting to scientific notation."""
    device = _deep_get(data, "%device_id", "device") or ""
    device = str(device).strip().upper()
    if device:
        return "\'" + device
    return "UNKNOWN"


def parse_single_json(json_path: str) -> Optional[Dict]:
    """
    Parse one MT7 GRR JSON log file.

    Expected format (v2.0):
      {
        "%device_id": { "device": "000016e8" },
        "ModelName": {
          "test": [
            { "pnum": "4024", "reference": "LED_RED_D303_INTENSITY",
              "actual": "30116", "status": "pass", "time": "..." },
            ...
          ]
        }
      }

    Returns a flat dict: one row per JSON file, columns = all test reference names.
    """
    # Try common encodings (UTF-8 first, then GBK for Chinese Windows, then Latin-1)
    data = None
    for enc in ("utf-8", "gbk", "gb2312", "latin-1"):
        try:
            with open(json_path, "r", encoding=enc) as f:
                data = json.load(f)
            break
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    if data is None:
        logger.warning(f"Failed to parse {json_path}: no valid encoding found")
        return None

    if not isinstance(data, dict):
        logger.warning(f"Unexpected JSON root type in {json_path}: {type(data)}")
        return None

    # Find the test array - skip meta keys (those starting with "%")
    test_array = None
    for key, val in data.items():
        if key.startswith("%"):
            continue
        if isinstance(val, dict) and "test" in val:
            test_array = val["test"]
            break

    if not test_array:
        logger.warning(f"No test array found in {json_path}")
        return None

    # Build flat record indexed by test reference name
    # Also extract LSL/USL if present in the JSON entry
    record = {}
    for entry in test_array:
        if not isinstance(entry, dict):
            continue
        ref  = str(_deep_get(entry, "reference") or "").strip()
        actual = _deep_get(entry, "actual")
        if ref:
            record[ref] = actual
            # Extract spec limits from JSON (MT7 stores as "min" / "max")
            min_val = _deep_get(entry, "min")
            max_val = _deep_get(entry, "max")
            if min_val is not None:
                record[f"{ref}_lsl"] = min_val
            if max_val is not None:
                record[f"{ref}_usl"] = max_val

    # Add metadata
    record["sn"]         = _extract_sn(data)
    record["appraiser"]  = _extract_appraiser(test_array)
    # Overall status from OVERALL_TEST_RESULT entry
    for e in test_array:
        if _deep_get(e, "reference") == "OVERALL_TEST_RESULT":
            record["status"] = str(_deep_get(e, "actual") or "UNKNOWN")
            break
    else:
        record["status"] = "UNKNOWN"
    # starttime from first test entry (used by group_into_grr_trials for sorting)
    if test_array and isinstance(test_array[0], dict):
        record["starttime"] = str(_deep_get(test_array[0], "time") or "")

    # Rename %device_id fields
    device_id = data.get("%device_id", {})
    if isinstance(device_id, dict):
        for k, v in device_id.items():
            if k != "entry_date" and isinstance(v, (str, int, float)):
                record[k] = v

    return record


def parse_json_folder(folder_path: str, recursive: bool = True) -> pd.DataFrame:
    """
    Parse all .json files in a folder into a single GRR DataFrame.

    Args:
        folder_path: Path to GRR Jason logs folder
        recursive:   Whether to search sub-folders

    Returns:
        DataFrame with all GRR records, columns = GRR_COLUMNS
    """
    folder = Path(folder_path)
    pattern = "**/*.json" if recursive else "*.json"
    json_files = sorted(folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)

    if not json_files:
        raise FileNotFoundError(f"No JSON files found in {folder_path}")

    logger.info(f"Found {len(json_files)} JSON files in {folder_path}")

    records = []
    for jf in json_files:
        rec = parse_single_json(str(jf))
        if rec:
            records.append(rec)

    if not records:
        raise ValueError(f"No valid records extracted from {folder_path}")

    df = pd.DataFrame(records)

    # Keep ALL columns from JSON (including dynamic test items like LED_XXX_INTENSITY)
    # Deduplicate by sn + appraiser + starttime
    df.drop_duplicates(
        subset=["sn", "appraiser", "starttime"],
        keep="first",
        inplace=True
    )
    df.reset_index(drop=True, inplace=True)

    logger.info(f"Parsed {len(df)} records, {len(df.columns)} columns")
    return df


def group_into_grr_trials(df: pd.DataFrame, parts: int = 10, operators: int = 3,
                           trials: int = 3) -> pd.DataFrame:
    """
    Re-index records into GRR trial structure.

    The MT7 tester runs each DUT multiple times by each operator.
    This function assigns:
      - trial number (1..trials)
      - part number (1..parts)  [sorted by sn]

    Args:
        df: DataFrame from parse_json_folder
        parts:    Number of DUTs (default 10 for GRR)
        operators: Number of operators/appraisers
        trials:   Measurement repetitions per (part, operator)

    Returns:
        DataFrame with added columns: trial, part_num
    """
    df = df.copy()

    # Get unique SNs, sorted
    unique_sns = df["sn"].dropna().unique()
    if len(unique_sns) < parts:
        logger.warning(f"Only {len(unique_sns)} unique SNs found, expected {parts}")
        parts = len(unique_sns)

    sorted_sns = sorted(unique_sns, key=lambda x: str(x))
    sn_to_part = {sn: i + 1 for i, sn in enumerate(sorted_sns)}
    df["part_num"] = df["sn"].map(sn_to_part)

    # Get unique appraisers
    appraisers = df["appraiser"].dropna().unique().tolist()
    if len(appraisers) < operators:
        logger.warning(f"Only {len(appraisers)} appraisers found, expected {operators}")
        operators = len(appraisers)

    df["trial"] = 1  # default

    # For each (part, appraiser), assign trial numbers sequentially
    def _assign_trials(group):
        group = group.sort_values("starttime")
        group["trial"] = [i % trials + 1 for i in range(len(group))]
        return group

    df = df.groupby(["part_num", "appraiser"], group_keys=False).apply(_assign_trials)
    df.reset_index(drop=True, inplace=True)

    return df


def export_to_excel(df: pd.DataFrame, output_path: str,
                    sheet_name: str = "log_YYYYMMDDHHMMSS") -> str:
    """
    Export the parsed GRR DataFrame to HVTE-M600099_GRR_data.xlsx format.

    Args:
        df:         Parsed GRR DataFrame
        output_path: Output .xlsx path
        sheet_name: Sheet name (default embeds current timestamp)

    Returns:
        Path to the saved file
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    sheet = sheet_name.replace("YYYYMMDDHHMMSS", timestamp)

    df_out = df.copy()

    # Drop LSL/USL auxiliary columns before writing (only needed for CPK computation in memory)
    lsl_usl_cols = [c for c in df_out.columns if c.endswith("_lsl") or c.endswith("_usl")]
    if lsl_usl_cols:
        logger.info(f"export_to_excel: dropping LSL/USL cols {lsl_usl_cols}")
        df_out.drop(columns=lsl_usl_cols, inplace=True)

    # Sort: primary = starttime (chronological order of measurement
    # capture); fall back to part_num / appraiser / trial as tie-breakers
    # so the ordering is stable when starttime is missing or identical.
    sort_cols = ["starttime", "part_num", "appraiser", "trial"]
    df_out.sort_values(
        by=[c for c in sort_cols if c in df_out.columns],
        inplace=True
    )
    df_out.reset_index(drop=True, inplace=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name=sheet, index=False)

        ws = writer.sheets[sheet]

        # Style header row
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Auto-width for data columns (limit to 50)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                min(max(max_len + 2, 10), 40)

        ws.freeze_panes = "A2"

    logger.info(f"Exported to {output_path}")
    return output_path


class JsonParser:
    """
    High-level interface: folder of JSON logs → intermediate GRR Excel.

    Usage:
        parser = JsonParser("/path/to/GRR_folder")
        parser.parse()                              # parse all JSON files
        parser.assign_trials(parts=10, ops=3, trials=3)  # assign GRR structure
        parser.export("raw_data.xlsx")  # write intermediate Excel
    """

    def __init__(self, folder: str):
        self.folder = Path(folder)
        self.df_raw: Optional[pd.DataFrame] = None
        self.df_grr: Optional[pd.DataFrame] = None

    def parse(self) -> "JsonParser":
        self.folder_path = self.combine_control_engine_logs(self.folder)

        # self.df_raw = parse_json_folder(str(self.folder))
        self.df_raw = parse_json_folder(self.folder_path)
        return self
    def combine_control_engine_logs(self, folder_path):
        folder_count = sum(1 for f in folder_path.iterdir() if f.is_dir())
        logger.info("folder_count is {}, type is {}".format(folder_count, type(folder_count)))
        if folder_count == 0:
            logger.info("no extra folder")
            return str(folder_path)
        elif folder_count == 2:
            folder1 = os.listdir(folder_path)[0]
            folder2 = os.listdir(folder_path)[1]
            dir1 = folder_path / folder1
            dir2 = folder_path / folder2
            # BASE_DIR / "templates"
            output_dir = OUTPUT_DIR / "combine_json_logs"
            output_dir.mkdir(parents=True, exist_ok=True)
            # 获取并排序（保证顺序一致！）
            for filename in os.listdir(str(output_dir)):
                file_path = os.path.join(str(output_dir), filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            logger.info("folder_count is {}, type is {}".format(dir2, dir1))
            files1 = sorted([f for f in dir1.glob("*.json")])
            files2 = sorted([f for f in dir2.glob("*.json")])
            if len(files1) != len(files2):
                raise ValueError("两个文件夹JSON数量不一致！")
            for f1, f2 in zip(files1, files2):
                with open(f1, "r", encoding="utf-8") as fp1, \
                        open(f2, "r", encoding="utf-8") as fp2:
                    data1 = json.load(fp1)
                    data2 = json.load(fp2)
                    # merged = self.merge_dict(data1, data2)
                data1[list(data1)[-1]]['test'] = data1[list(data1)[-1]]['test'][:-1] + data2[list(data2)[-1]]['test']
                print(data1)
                out_path = output_dir / f1.name
                with open(out_path, "w", encoding="utf-8") as out:
                    json.dump(data1, out, indent=4, ensure_ascii=False)
            print("✅ 合并完成，共处理:", len(files1))
            return str(output_dir)
        else:
            raise ValueError("log 文件夹有两个以上的文件夹")

    def get_allt_test_infors(self):
        self.dirs = os.listdir("{}".format(self.folder_path))
        for sub_dir in self.dirs:
            with open(r"{}\\{}".format(self.folder_path, sub_dir), 'r', encoding="utf-8") as f:
                json_content = json.load(f)
                break
        for key in json_content:
            for sub_key in json_content[key]:
                if sub_key == "test":
                    self.test_data = json_content[key]["test"]
                    print(self.test_data)
                    break
        test_data_file = os.path.join(
            OUTPUT_DIR,
            "test_data.json",
        )
        data = json.dumps(self.test_data, indent=2)
        with open(test_data_file, "w", encoding="utf-8") as f:
            f.write(data)
            f.close()
        return True

    def assign_trials(self, parts: int = 10, operators: int = 3,
                      trials: int = 3) -> "JsonParser":
        if self.df_raw is None:
            raise RuntimeError("Call parse() first")
        self.df_grr = group_into_grr_trials(
            self.df_raw, parts=parts, operators=operators, trials=trials
        )
        return self

    def export(self, output_path: str,
               sheet_name: str = "log_YYYYMMDDHHMMSS") -> str:
        df = self.df_grr if self.df_grr is not None else self.df_raw
        if df is None:
            raise RuntimeError("No data to export")
        excel_path = export_to_excel(df, output_path, sheet_name)
        # Write LSL/USL specs to a sidecar JSON so CPK can read them after Excel reload
        self._write_specs_sidecar(output_path)
        self.get_allt_test_infors()
        return excel_path

    def _write_specs_sidecar(self, excel_path: str):
        """Save LSL/USL columns to a .specs.json next to the Excel file."""
        import json
        df = self.df_grr if self.df_grr is not None else self.df_raw
        if df is None:
            return
        lsl_usl = {}
        for c in df.columns:
            if c.endswith("_lsl") or c.endswith("_usl"):
                vals = df[c].dropna()
                val = vals.iloc[0] if len(vals) > 0 else None
                # Convert numpy/pandas types to native Python for JSON serialization
                if hasattr(val, 'item'):
                    val = val.item()
                elif hasattr(val, 'to_pydatetime'):
                    val = str(val)
                lsl_usl[c] = val
        if not lsl_usl:
            return
        sidecar = Path(excel_path).parent / (Path(excel_path).name + ".specs.json")
        with open(sidecar, "w") as f:
            json.dump(lsl_usl, f)
        logger.info(f"Specs sidecar written: {sidecar}")

    def get_df(self) -> Optional[pd.DataFrame]:
        return self.df_grr if self.df_grr is not None else self.df_raw
