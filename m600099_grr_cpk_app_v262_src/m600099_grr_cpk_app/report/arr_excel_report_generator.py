# ──────────────────────────────────────────────
#  AR&R Excel Report Generator
#  Generates Form-004091 AR&R report from parsed Jason logs
# ──────────────────────────────────────────────
import logging
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Patch MergedCell.__setattr__ so .value writes are silently ignored ──
# openpyxl raises AttributeError when you try to set .value on a MergedCell
# (any cell in a merged range that is not the top-left cell).
# This patch makes those writes a no-op so the rest of the code is unaffected.
import openpyxl.worksheet.merge as _mrg
from utils.config import ARR_LAYOUT, ARR_FORM_DEFAULTS
from core.json_parser import JsonParser
__orig_setattr = _mrg.MergedCell.__setattr__
def _safe_setattr(self, name, val):
    if name == "value":
        return
    __orig_setattr(self, name, val)
_mrg.MergedCell.__setattr__ = _safe_setattr


logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────
CLR_HEADER_BG  = "1F4E79"
CLR_HEADER_FG  = "FFFFFF"
CLR_SUBHDR_BG  = "2E75B6"
CLR_SUBHDR_FG  = "FFFFFF"
CLR_GRN        = "00B050"  # dark green (reference standard)
CLR_YLW        = "FFFF00"  # yellow (reference standard)
CLR_RED        = "FF0000"  # red (reference standard)
CLR_ALT        = "EBF3FB"
CLR_PASS_FG    = "375623"
CLR_FAIL_FG    = "9C0006"

def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)

def _thin() -> Border:
    # Use FF prefix (ARGB) so openpyxl stores opaque color, not transparent
    s = Side(style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG,
         size=10, bold=True, wrap=False, align="center"):
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill  = _fill(bg)
    cell.border = _thin()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def _dat(cell, value, bg=None, fg="000000", bold=False,
         fmt=None, align="center", border=True):
    """Write value + style into cell. border=True preserves template borders."""
    cell.value = value
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=10)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border    = _thin()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


def _val(cell, value, fmt=None):
    """Set value + number_format only — preserves template font/fill/border."""
    cell.value = value
    if fmt:
        cell.number_format = fmt


# ──────────────────────────────────────────────
#  Attribute Agreement Analysis
# ──────────────────────────────────────────────

def compute_attribute_agreement(
    df_sheet1: pd.DataFrame,
    df_summary: pd.DataFrame
) -> Dict:
    """
    Compute attribute agreement metrics.

    Sheet1 columns: Column (expected), Samples, Appraisers, attribute (actual)
    Summary columns: Item, Error Code, Standard, A-1..C-3

    Returns dict:
      - overall_eff      : overall effectiveness %
      - kappa            : Cohen's Kappa
      - by_appraiser     : {appraiser: effectiveness%}
      - within_appraiser : {appraiser: within-appraiser eff%}
      - vs_std           : {appraiser: vs-standard eff%}
      - between_appraiser: between-appraiser eff%
    """
    n = len(df_sheet1)
    if n == 0:
        return {}

    # ── Overall effectiveness ────────────────────
    correct = (df_sheet1["Column"] == df_sheet1["attribute"]).sum()
    overall_eff = correct / n * 100 if n > 0 else 0

    # ── Cohen's Kappa ────────────────────────────
    po = overall_eff / 100  # observed agreement
    # Expected agreement by chance
    n1 = df_sheet1["Column"].sum()          # expected positive
    n2 = n - n1                              # expected negative
    m1 = df_sheet1["attribute"].sum()        # actual positive
    m2 = n - m1                              # actual negative
    pe = (n1 * m1 + n2 * m2) / (n * n) if n > 0 else 0
    kappa = (po - pe) / (1 - pe) if (1 - pe) != 0 else 0

    # ── Per appraiser vs standard ───────────────
    by_app = {}
    within_app = {}
    appraisers = sorted(df_sheet1["Appraisers"].unique())

    for app in appraisers:
        sub = df_sheet1[df_sheet1["Appraisers"] == app]
        nc  = (sub["Column"] == sub["attribute"]).sum()
        by_app[app] = nc / len(sub) * 100 if len(sub) > 0 else 0

    # ── Within appraiser agreement ──────────────
    # For each appraiser, compare repeated trials on same sample
    for app in appraisers:
        sub = df_sheet1[df_sheet1["Appraisers"] == app].copy()
        sub_sorted = sub.sort_values(["Samples"])
        # Group by sample — count agreement between trials
        # If all trials for a sample agree → correct; else → incorrect
        grp_agree = sub_sorted.groupby("Samples")["attribute"].apply(
            lambda x: 1 if x.nunique() == 1 else 0
        )
        within_app[app] = grp_agree.sum() / len(grp_agree) * 100 if len(grp_agree) > 0 else 0

    # ── Between appraiser agreement ─────────────
    # For each sample, check if all appraisers agree on the result
    sample_means = df_sheet1.groupby(["Samples", "Appraisers"])["attribute"].mean()
    sample_std = df_sheet1.groupby("Samples")["attribute"].std(ddof=0).fillna(0)
    between_count = (sample_std == 0).sum()
    n_samples = df_sheet1["Samples"].nunique()
    between_eff = between_count / n_samples * 100 if n_samples > 0 else 0

    return {
        "overall_eff":       overall_eff,
        "kappa":             kappa,
        "by_appraiser":      by_app,
        "within_appraiser":  within_app,
        "vs_standard":       by_app,   # same as vs standard
        "between_appraiser": between_eff,
    }


# ──────────────────────────────────────────────
#  AR&R Report Generator
# ──────────────────────────────────────────────

class ARRReportGenerator:
    """
    Generates Form-004091 AR&R Excel report.

    Workflow:
      1. Load AR&R template Excel (or generate blank template)
      2. Parse Jason logs → df_appraisers (sn, appraiser, test_item → result)
      3. Map samples to error codes from Summary sheet
      4. Compute attribute agreement metrics
      5. Fill all sheets
    """

    def __init__(self, template_path: Optional[str] = str(Path(__file__).parent / "AR_R_Template.xlsx"), output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.template_path = template_path

    # ── Internal helpers ─────────────────────────

    @staticmethod
    def _load_template(path: str) -> openpyxl.Workbook:
        """Load AR&R template workbook, preserve formatting."""
        return openpyxl.load_workbook(path)

    def _ensure_template(self) -> openpyxl.Workbook:
        if self.template_path and Path(self.template_path).exists():
            return self._load_template(self.template_path)
        return self._create_blank_template()

    def _create_blank_template(self) -> openpyxl.Workbook:
        """Create a blank Form-004091 workbook from scratch."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Cover Page ──────────────────────────────
        ws = wb.create_sheet("Cover Page", 0)
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 12

        ws.merge_cells("B2:H2")
        _hdr(ws.cell(2, 2), "APPROVALS", bg=CLR_HEADER_BG, size=10)
        ws.row_dimensions[2].height = 20

        headers = ["DEPT", "NAME", "TITLE", "DATE"]
        for j, h in enumerate(headers, 2):
            _hdr(ws.cell(3, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=9)
        _hdr(ws.cell(3, 6), "AR and R Form(DMD)", bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=9)

        ws.merge_cells("F6:G6")
        _hdr(ws.cell(6, 6), "TEMPLATE NUMBER\nFORM-004091", bg="F5F5F5", fg="000000",
             size=9, bold=False, align="center")
        ws.merge_cells("H6:H6")
        _hdr(ws.cell(6, 8), "REVISION\nB", bg="F5F5F5", fg="000000",
             size=9, bold=False, align="center")

        # Revision history
        ws.merge_cells("B10:H10")
        _hdr(ws.cell(10, 2), "REVISION HISTORY", bg=CLR_HEADER_BG, size=9)
        rev_hdrs = ["REVISION", "DESCRIPTION OF CHANGE", "", "", "ORIGINATOR", "RELEASE DATE", "", ""]
        for j, h in enumerate(rev_hdrs, 2):
            if h:
                _hdr(ws.cell(12, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=8)
        for row, rev, desc, orig, date in [
            (13, "A", "Transfer from QDMS to EQMS", "Li Rao", "In EQMS"),
            (14, "B", "Update header and footer for the template", "Li Rao", "In EQMS"),
        ]:
            ws.cell(row, 2).value = rev
            ws.merge_cells(f"C{row}:E{row}")
            ws.cell(row, 3).value = desc
            ws.cell(row, 6).value = orig
            ws.cell(row, 7).value = date
            for c in [ws.cell(row, 2), ws.cell(row, 3), ws.cell(row, 6), ws.cell(row, 7)]:
                c.border = _thin()

        ws.merge_cells("B20:H20")
        ws.cell(20, 2).value = ("Note: The cover page is only used for electronic file "
                                 "traceability of change history, no need print out "
                                 "the cover page when do hardcopy record")
        ws.cell(20, 2).font = Font(name="Calibri", italic=True, size=8, color="595959")

        # ══════════════════════════════════════════════════════════════════
        #  AR&R Form  – pixel-perfect match to reference template
        #  Font: Arial throughout (except as noted)
        #  Fills: transparent=none, white=00000000
        # ══════════════════════════════════════════════════════════════════
        ws = wb.create_sheet("AR&R Form", 1)

        # ── Column widths ──────────────────────────────────────────
        ws.column_dimensions["A"].width = 7.23
        ws.column_dimensions["B"].width = 14.0
        ws.column_dimensions["C"].width = 7.69
        ws.column_dimensions["D"].width = 7.84
        ws.column_dimensions["E"].width = 8.0
        ws.column_dimensions["F"].width = 5.0
        ws.column_dimensions["G"].width = 14.0
        ws.column_dimensions["H"].width = 7.23
        ws.column_dimensions["I"].width = 16.0
        ws.column_dimensions["J"].width = 7.23
        ws.column_dimensions["K"].width = 12.0

        def _cell_fill_none(cell):
            """Transparent (no pattern fill)."""
            cell.fill = PatternFill(fill_type=None)

        def _cell_fill_white(cell):
            """White fill (explicit)."""
            cell.fill = PatternFill("solid", fgColor="FFFFFF")

        def _meta_left(ws, row, label_text, row2_top_border=False):
            """Left group: A:B = label, C:D = value.  Fills: transparent."""
            ws.row_dimensions[row].height = 18.75
            # A:B merged = label
            ws.merge_cells(f"A{row}:B{row}")
            ca = ws.cell(row, 1)
            ca.value = label_text
            ca.font = Font(name="Arial", size=10)
            ca.border = Border(
                left=Side(style="medium"),
                top=Side(style="medium") if row2_top_border else Side(style=None),
                bottom=Side(style=None))
            _cell_fill_none(ca)
            # C:D merged = value
            ws.merge_cells(f"C{row}:D{row}")
            cc = ws.cell(row, 3)
            cc.font = Font(name="Arial", size=10)
            cc.alignment = Alignment(horizontal="left", vertical="top")
            if row2_top_border:
                cc.border = Border(top=Side(style="medium"), bottom=Side(style="thin"))
            else:
                cc.border = Border()   # no border on non-row-2 metadata value cells
            _cell_fill_none(cc)

        def _meta_right(ws, row, label_text, value_text="", row2_top_border=False):
            """Right group: G:H = label, I:J = value."""
            ws.row_dimensions[row].height = 18.75
            # G:H merged = label
            ws.merge_cells(f"G{row}:H{row}")
            cg = ws.cell(row, 7)
            cg.value = label_text
            cg.font = Font(name="Arial", size=10)
            cg.alignment = Alignment(horizontal="left")
            cg.border = Border(top=Side(style="medium") if row2_top_border else Side(style=None))
            _cell_fill_none(cg)
            # I:J merged = value
            ws.merge_cells(f"I{row}:J{row}")
            ci = ws.cell(row, 9)
            ci.value = value_text
            ci.font = Font(name="Arial", size=10)
            ci.alignment = Alignment(horizontal="left", vertical="top")
            if row2_top_border:
                ci.border = Border(top=Side(style="medium"), bottom=Side(style="thin"))
            else:
                ci.border = Border()   # no border on non-row-2 value cells
            _cell_fill_none(ci)
            # K: medium right border (row 2 only), no fill
            ck = ws.cell(row, 11)
            ck.border = Border(
                right=Side(style="medium"),
                top=Side(style="medium") if row2_top_border else Side(style=None))
            _cell_fill_none(ck)

        # ── Row 1: title ────────────────────────────────────────────
        ws.row_dimensions[1].height = 30.75
        ws.merge_cells("A1:K1")
        c1 = ws.cell(1, 1)
        c1.value = "AR&R Form"
        c1.font = Font(name="Arial", size=14, bold=True)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        _cell_fill_none(c1)
        # No borders on title row

        # ── Rows 2-7: metadata ───────────────────────────────────────
        meta_left = [
            (2, "Part Number :",    True),
            (3, "Instrument name:", False),
            (4, "Department Name:", False),
            (5, "Inspector Number:", False),
            (6, "Inspector Number:", False),
            (7, "Inspector Number:", False),
        ]
        meta_right = [
            (2, "Description :",    "Alpha Finished ",  True),
            (3, "Instrument No:",    "HVTE-M600271",    False),
            (4, "Reported by :",    "Bin Liao(DMN)",    False),
            (5, "Fixture No : ",     "HVTE-M600271",    False),
            (6, "Project Name : ",   "Alpha",           False),
            (7, "Date Inspected:",  "Apr 08,2026",     False),
        ]
        for row, label, row2_flag in meta_left:
            _meta_left(ws, row, label, row2_top_border=row2_flag)
        for row, label, val, row2_flag in meta_right:
            _meta_right(ws, row, label, val, row2_top_border=row2_flag)
        # Fill K2 right medium border for row 2
        ws.cell(2, 11).border = Border(right=Side(style="medium"), top=Side(style="medium"))

        # ── Row 8: thin separator ───────────────────────────────────
        ws.row_dimensions[8].height = 5.25
        for c_idx in range(1, 12):
            cell = ws.cell(8, c_idx)
            cell.border = Border(bottom=Side(style="thin"))
            _cell_fill_none(cell)

        # ── Row 9: "Parameter" label ────────────────────────────────
        ws.row_dimensions[9].height = 16.0
        # A9: left medium, transparent fill
        ca9 = ws.cell(9, 1)
        ca9.border = Border(left=Side(style="medium"))
        _cell_fill_none(ca9)
        # B9:K9 merged = "Parameter"
        ws.merge_cells("B9:K9")
        cb9 = ws.cell(9, 2)
        cb9.value = "Parameter"
        cb9.font = Font(name="Arial", size=10)
        cb9.alignment = Alignment(horizontal="center")
        _cell_fill_white(cb9)
        # Row 9 borders:
        # A9 = standalone, left-medium/right-thin
        # B9:K9 = merged (Parameter label); set border on anchor B9 only
        # K9 = right-medium (inherits via merged range)
        ca9 = ws.cell(9, 1)
        ca9.border = Border(left=Side(style="medium"), right=Side(style="thin"),
                             top=Side(style="medium"), bottom=Side(style="medium"))
        cb9 = ws.cell(9, 2)    # B9 is the merged-range anchor; border set here applies to whole range
        cb9.border = Border(top=Side(style="medium"), bottom=Side(style="medium"))

        # ── Row 10: column headers ──────────────────────────────────
        ws.row_dimensions[10].height = 16.0
        hdrs = ["Sample","A-1","A-2","A-3","B-1","B-2","B-3","C-1","C-2","C-3","Basis"]
        for c_idx, h in enumerate(hdrs, 1):
            cell = ws.cell(10, c_idx)
            cell.value = h
            is_A = (c_idx == 1)
            is_K = (c_idx == 11)
            is_data = (1 < c_idx < 11)
            cell.font = Font(name="Arial", size=9 if is_data else 10)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=(c_idx > 1))
            cell.border = Border(
                left=Side(style="medium") if is_A else Side(style="thin"),
                right=Side(style="medium") if is_K else Side(style="thin"),
                top=Side(style="medium"),
                bottom=Side(style="thin"))
            _cell_fill_white(cell)

        # ── Rows 11-20: data cells ─────────────────────────────────
        for r in range(11, 21):
            ws.cell(r, 1).value = r - 10
            ws.cell(r, 1).font = Font(name="Arial", size=10)
            ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
            # A column: left=medium, right/top/bottom=thin
            ac = ws.cell(r, 1)
            ac.border = Border(left=Side(style="medium"), right=Side(style="thin"),
                                top=Side(style="thin"), bottom=Side(style="thin"))
            _cell_fill_white(ac)
            for c_idx in range(2, 12):
                is_K = (c_idx == 11)
                cell = ws.cell(r, c_idx)
                # B-J: all thin; K: transparent fill, right=medium
                # Must use new Border() object (openpyxl Border is immutable)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="medium") if is_K else Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"))
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_K:
                    _cell_fill_none(cell)   # K: transparent (matching reference)
                else:
                    _cell_fill_white(cell)  # B-J: white

        # ── Row 21: Remarks ─────────────────────────────────────────
        ws.row_dimensions[21].height = 25.5
        ws.merge_cells("A21:K21")
        rm = ws.cell(21, 1)
        rm.value = 'Remarks:" 1 " means an acceptable decision; "0" means an unacceptable decision.'
        rm.font = Font(name="Arial", size=12)
        rm.alignment = Alignment(horizontal="left")
        # Set full border (Row 21: left/right/bottom = medium, top = none)
        rm.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                           top=Side(style=None), bottom=Side(style="medium"))
        _cell_fill_none(rm)
        # Fill A21:K21 with transparent (overridden by merge anchor A21)
        for c_idx in range(1, 12):
            cell = ws.cell(21, c_idx)
            _cell_fill_none(cell)

        # ── Row 22: Template link ───────────────────────────────────
        ws.row_dimensions[22].height = 18.0
        ws.merge_cells("A22:K22")
        tl = ws.cell(22, 1)
        tl.value = "This Template Link to #WI-002388"
        tl.font = Font(name="Arial", size=12)
        _cell_fill_none(tl)

        # ══════════════════════════════════════════════════════════════════
        #  AR&R Report  – pixel-perfect match to reference template
        # ══════════════════════════════════════════════════════════════════
        ws = wb.create_sheet("AR&R Report", 2)

        # ── Column widths ──────────────────────────────────────────
        ws.column_dimensions["A"].width = 1.23
        ws.column_dimensions["B"].width = 7.07
        ws.column_dimensions["K"].width = 9.23
        ws.column_dimensions["M"].width = 7.07
        ws.column_dimensions["Q"].width = 8.84
        for col in ["C","D","E","F","G","H","I","J","L","N","O","P"]:
            ws.column_dimensions[col].width = 7.0

        # ── Row 1: thin top separator ──────────────────────────────
        ws.row_dimensions[1].height = 5.25
        for c_idx in range(1, 17):
            ws.cell(1, c_idx).border = Border(top=Side(style="thin"))
            _cell_fill_none(ws.cell(1, c_idx))

        # ── Row 2: "1.AR&R Data" ──────────────────────────────────
        ws.row_dimensions[2].height = 16.0
        ws.merge_cells("B2:P2")
        c2 = ws.cell(2, 2)
        c2.value = "1.AR&R Data"
        c2.font = Font(name="Arial", size=12)
        c2.alignment = Alignment(horizontal="center")
        c2.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"))
        _cell_fill_none(c2)

        # ── Rows 3-13: chart area ─────────────────────────────────
        for r in range(3, 14):
            ws.row_dimensions[r].height = 20.25 if r in [7, 9] else 16.0
            for c_idx in range(2, 17):
                ws.cell(r, c_idx).border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
                _cell_fill_none(ws.cell(r, c_idx))

        # ── Row 14: "2.Graph Report" | "3.Conclusion:" ─────────────
        ws.row_dimensions[14].height = 16.0
        ws.merge_cells("B14:I14")
        c14a = ws.cell(14, 2)
        c14a.value = "2.Graph Report"
        c14a.font = Font(name="Arial", size=12)
        c14a.alignment = Alignment(horizontal="center")
        c14a.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"))
        _cell_fill_none(c14a)

        ws.merge_cells("J14:P14")
        c14b = ws.cell(14, 10)
        c14b.value = "3.Conclusion:"
        c14b.font = Font(name="Arial", size=12)
        c14b.alignment = Alignment(horizontal="center")
        c14b.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"))
        _cell_fill_none(c14b)

        # ── Row 15: Inspector header + A/B/C ────────────────────────
        ws.row_dimensions[15].height = 16.0
        ws.merge_cells("J15:L15")
        h15 = ws.cell(15, 10)
        h15.value = "Inspector"
        h15.font = Font(name="Arial", size=12)
        h15.alignment = Alignment(horizontal="center")
        h15.border = Border(
            left=Side(style="medium"), right=Side(style="thin"),
            bottom=Side(style="thin"))
        _cell_fill_none(h15)
        for c_idx, lbl in [(13,"A"),(14,"B"),(15,"C")]:
            cell = ws.cell(15, c_idx)
            cell.value = lbl
            cell.font = Font(name="Arial", size=12)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                bottom=Side(style="thin"))
            _cell_fill_none(cell)

        # ── Rows 16-19: Conclusion table ───────────────────────────
        conclusion_defs = [
            (16, "Within Appraisers \n检验员自身一致性",        39.5),
            (17, "Each Appraiser vs Standard\n每个检验员与标准一致性 ", 41.0),
            (18, "Between Appraisers\n检验员之间一致性",         32.0),
            (19, "All Appraisers vs Standard \n所有检验员与标准一致性", 32.25),
        ]
        num_syms = ["①","②","③","④"]
        for idx, (row_num, desc, ht) in enumerate(conclusion_defs):
            ws.row_dimensions[row_num].height = ht
            # col J: number (Meiryo, transparent fill)
            jc = ws.cell(row_num, 10)
            jc.value = num_syms[idx]
            jc.font = Font(name="Meiryo", size=12)
            jc.alignment = Alignment(horizontal="center")
            jc.border = Border(
                left=Side(style="medium"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            _cell_fill_none(jc)
            # cols K:L: description (merged, Arial 8, transparent)
            ws.merge_cells(f"K{row_num}:L{row_num}")
            kc = ws.cell(row_num, 11)
            kc.value = desc
            kc.font = Font(name="Arial", size=8)
            kc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            kc.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                               top=Side(style="thin"), bottom=Side(style="thin"))
            _cell_fill_none(kc)
            # cols M,N,O: A,B,C result cells
            for c_idx in [13, 14, 15]:
                cc = ws.cell(row_num, c_idx)
                cc.font = Font(name="Arial", size=10 if row_num < 18 else 12)
                cc.alignment = Alignment(horizontal="center",
                                        vertical="center" if row_num < 18 else None)
                cc.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                   top=Side(style="thin"), bottom=Side(style="thin"))
                _cell_fill_none(cc)

        # rows 18-19: M18:O18 and M19:O19 merged (only col M filled)
        ws.merge_cells("M18:O18")
        ws.merge_cells("M19:O19")

        # ── Row 20: "4.Notes:" label ──────────────────────────────
        ws.row_dimensions[20].height = 52.5
        j20 = ws.cell(20, 10)
        j20.value = "4.Notes:"
        j20.font = Font(name="Arial", size=12)
        j20.border = Border(left=Side(style="medium"))
        _cell_fill_none(j20)

        # ── Rows 21-23: Notes text (J21:P23 merged) ──────────────
        ws.row_dimensions[21].height = 38.25
        ws.row_dimensions[22].height = 15.0
        ws.row_dimensions[23].height = 16.0
        ws.merge_cells("J21:P23")
        n21 = ws.cell(21, 10)
        n21.value = (
            "1) %AR&R≥90%，Good \n"
            "2) 80%≤%AR&R<90%，Acceptable.\n"
            "3) %AR&R＜80%，Bad."
        )
        n21.font = Font(name="Arial", size=12)
        n21.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n21.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="thin"), bottom=Side(style="medium"))
        _cell_fill_none(n21)

        # ── Row 24: Template link ──────────────────────────────────
        ws.row_dimensions[24].height = 16.0
        ws.merge_cells("B24:P24")
        tl24 = ws.cell(24, 2)
        tl24.value = "This Template Link to #WI-002388"
        tl24.font = Font(name="Arial", size=12)
        _cell_fill_none(tl24)

        # ══════════════════════════════════════════════════════
        #  AR&R Report  – matches sample template exactly
        # ══════════════════════════════════════════════════════
        ws = wb.create_sheet("AR&R Report", 2)

        # ── Column widths ─────────────────────────
        ws.column_dimensions["A"].width = 1.23
        ws.column_dimensions["B"].width = 7.07
        ws.column_dimensions["K"].width = 9.23
        ws.column_dimensions["M"].width = 7.07
        ws.column_dimensions["Q"].width = 8.84
        for col in ["C","D","E","F","G","H","I","J","L","N","O","P"]:
            ws.column_dimensions[col].width = 7.0

        # ── Row 1: thin top separator ────────────
        ws.row_dimensions[1].height = 5.25
        for c in range(1, 17):
            ws.cell(1,c).border = Border(top=Side(style="thin"))

        # ── Row 2: "1.AR&R Data" section header ─
        ws.row_dimensions[2].height = 16.0
        ws.merge_cells("B2:P2")
        c2 = ws.cell(2, 2)
        c2.value = "1.AR&R Data"
        c2.font  = Font(name="Arial", size=12)
        c2.fill  = PatternFill("solid", fgColor="D9D9D9")
        c2.alignment = Alignment(horizontal="center")
        c2.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"))

        # ── Rows 3-13: chart area (thin borders) ─
        for r in range(3, 14):
            ws.row_dimensions[r].height = 20.25 if r in [7, 9] else 16.0
            for c in range(2, 17):
                ws.cell(r,c).border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

        # ── Row 14: "2.Graph Report" | "3.Conclusion:" ─
        ws.row_dimensions[14].height = 16.0
        ws.merge_cells("B14:I14")
        c14a = ws.cell(14, 2)
        c14a.value = "2.Graph Report"
        c14a.font  = Font(name="Arial", size=12)
        c14a.fill  = PatternFill("solid", fgColor="D9D9D9")
        c14a.alignment = Alignment(horizontal="center")
        c14a.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"))

        ws.merge_cells("J14:P14")
        c14b = ws.cell(14, 10)
        c14b.value = "3.Conclusion:"
        c14b.font  = Font(name="Arial", size=12)
        c14b.fill  = PatternFill("solid", fgColor="D9D9D9")
        c14b.alignment = Alignment(horizontal="center")
        c14b.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"))

        # ── Row 15: Inspector header + A/B/C cols ─
        ws.row_dimensions[15].height = 16.0
        ws.merge_cells("J15:L15")
        h15 = ws.cell(15, 10)
        h15.value = "Inspector"
        h15.font  = Font(name="Arial", size=12)
        h15.alignment = Alignment(horizontal="center")
        h15.border = Border(
            left=Side(style="medium"), right=Side(style="thin"),
            top=Side(style=None), bottom=Side(style="thin"))
        for c, lbl in [(13,"A"),(14,"B"),(15,"C")]:
            cell = ws.cell(15, c)
            cell.value = lbl
            cell.font  = Font(name="Arial", size=12)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style=None), bottom=Side(style="thin"))

        # ── Rows 16-19: Conclusion table ─────────
        conclusion_defs = [
            (16, "Within Appraisers \n检验员自身一致性",       39.5),
            (17, "Each Appraiser vs Standard\n每个检验员与标准一致性 ", 41.0),
            (18, "Between Appraisers\n检验员之间一致性",        32.0),
            (19, "All Appraisers vs Standard \n所有检验员与标准一致性", 32.25),
        ]
        num_syms = ["①","②","③","④"]
        for idx, (row_num, desc, ht) in enumerate(conclusion_defs):
            ws.row_dimensions[row_num].height = ht
            # col J: number (Meiryo)
            jc = ws.cell(row_num, 10)
            jc.value = num_syms[idx]
            jc.font  = Font(name="Meiryo", size=12)
            jc.alignment = Alignment(horizontal="center")
            jc.border = Border(
                left=Side(style="medium"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            # cols K:L: description (merged, Arial 8, wrap)
            ws.merge_cells(f"K{row_num}:L{row_num}")
            kc = ws.cell(row_num, 11)
            kc.value = desc
            kc.font  = Font(name="Arial", size=8)
            kc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            kc.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            # cols M,N,O: A,B,C result cells
            for c in [13, 14, 15]:
                cc = ws.cell(row_num, c)
                cc.font = Font(name="Arial", size=10 if row_num < 18 else 12)
                cc.alignment = Alignment(horizontal="center",
                                         vertical="center" if row_num < 18 else None)
                cc.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

        # rows 18-19: M18:O18 and M19:O19 merged (only col M filled)
        ws.merge_cells("M18:O18")
        ws.merge_cells("M19:O19")

        # ── Row 20: "4.Notes:" label ─────────────
        ws.row_dimensions[20].height = 52.5
        j20 = ws.cell(20, 10)
        j20.value = "4.Notes:"
        j20.font  = Font(name="Arial", size=12)
        j20.border = Border(left=Side(style="medium"))

        # ── Rows 21-23: Notes text merged J21:P23 ─
        ws.row_dimensions[21].height = 38.25
        ws.row_dimensions[22].height = 15.0
        ws.row_dimensions[23].height = 16.0
        ws.merge_cells("J21:P23")
        n21 = ws.cell(21, 10)
        n21.value = (
            "1) %AR&R≥90%，Good \n"
            "2) 80%≤%AR&R<90%，Acceptable.\n"
            "3) %AR&R＜80%，Bad."
        )
        n21.font = Font(name="Arial", size=12)
        n21.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n21.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="thin"), bottom=Side(style="medium"))

        # ── Row 24: Template link ─────────────────
        ws.row_dimensions[24].height = 16.0
        ws.merge_cells("B24:P24")
        tl24 = ws.cell(24, 2)
        tl24.value = "This Template Link to #WI-002388"
        tl24.font  = Font(name="Arial", size=12)

        # ── Summary ────────────────────────────────
        ws = wb.create_sheet("Summary", 3)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 12
        for j in range(4, 13):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 10

        ws.merge_cells("A1:M1")
        _hdr(ws.cell(1, 1), "Item", bg=CLR_HEADER_BG, size=10)
        ws.cell(1, 4).value = "Number of trial (MT7H Tester Functional Test System)"
        ws.cell(1, 4).font  = Font(name="Calibri", bold=True, size=10, color=CLR_HEADER_FG)
        ws.cell(1, 4).fill  = _fill(CLR_HEADER_BG)

        hrow = ["", "", ""] + ["A-1","A-2","A-3","B-1","B-2","B-3","C-1","C-2","C-3"]
        for j, h in enumerate(hrow, 1):
            _hdr(ws.cell(2, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=9)
        ws.row_dimensions[2].height = 20

        for r in range(3, 13):
            ws.cell(r, 1).border = _thin()
            ws.cell(r, 2).border = _thin()
            ws.cell(r, 3).border = _thin()
            for c in range(4, 13):
                ws.cell(r, c).border = _thin()
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, c).font = Font(name="Calibri", size=9)

        # ── Faults simulation ────────────────────────
        ws = wb.create_sheet("Faults simulation", 4)
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 35
        for j, h in enumerate(["No.", "Error code", "Description"], 1):
            _hdr(ws.cell(1, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=9)
        for r in range(2, 6):
            for c in range(1, 4):
                ws.cell(r, c).border = _thin()
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, c).font = Font(name="Calibri", size=9)

        # ── Sheet1 ─────────────────────────────────
        ws = wb.create_sheet("Sheet1", 5)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 12
        for j, h in enumerate(["Column", "Samples", "Appraisers", "attribute"], 1):
            _hdr(ws.cell(1, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=9)
        for r in range(2, 92):
            for c in range(1, 5):
                ws.cell(r, c).border = _thin()
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, c).font = Font(name="Calibri", size=9)

        return wb

    @staticmethod
    def _build_sheet1(
        df_jason: pd.DataFrame,
        sample_map: Dict[int, Dict],
        n_trials: int = 3,
        inspector_numbers: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """
        Build Sheet1 data from parsed Jason logs.

        df_jason columns (from parse_json_folder):
            appraiser, sample_num, trial, status, <test_items>

        Sheet1 columns: Column (expected), Samples, Appraisers, attribute (actual)
        Appraisers values: inspector_numbers[index] if available, else raw appraiser label.
        """
        # Defensive: ensure 'attribute' column exists (PASS→1, FAIL→0)
        if "attribute" not in df_jason.columns:
            df_jason = df_jason.copy()
            df_jason["attribute"] = df_jason["status"].apply(
                lambda s: 1 if str(s).lower().startswith("pass") else 0
            )

        rows = []
        appraisers = sorted(df_jason["appraiser"].dropna().unique().tolist())
        logger.info("_build_sheet1: inspector_numbers=%s  appraisers=%s", inspector_numbers, appraisers)

        # Build appraiser -> inspector number mapping
        if inspector_numbers:
            app_insp_map = {
                app: inspector_numbers[i] if i < len(inspector_numbers) else app
                for i, app in enumerate(appraisers)
            }
        else:
            app_insp_map = {app: app for app in appraisers}

        for appraiser in appraisers:
            app_df = df_jason[df_jason["appraiser"] == appraiser]

            # Build two lookups keyed by (trial_int, sample_int):
            #   status_lookup: for Column (pass/fail vs standard)
            #   attr_lookup:   for attribute (actual measurement value 0 or 1)
            status_lookup = {}
            attr_lookup   = {}
            for _, row in app_df.iterrows():
                sn    = int(row.get("sample_num", 0))
                tr    = int(row.get("trial", 1))
                stat  = str(row.get("status", "")).lower()
                attr_val = int(row.get("attribute", 1))  # raw measurement or derived from status
                if "attribute" not in row:
                    # Derive attribute from status: PASS=1, FAIL=0
                    stat = str(row.get("status", "")).lower()
                    attr_val = 1 if "pass" in stat else 0
                if pd.notna(row.get("sample_num")):
                    status_lookup[(tr, sn)] = 1 if "pass" in stat else 0
                    attr_lookup  [(tr, sn)] = attr_val

            trials  = sorted(app_df["trial"].dropna().unique(), key=lambda x: int(x))
            samples = sorted(app_df["sample_num"].dropna().unique(), key=lambda x: int(x))

            for trial in trials:
                for sample_num in samples:
                    t = int(trial); s = int(sample_num)
                    info     = sample_map.get(s, {})
                    expected = info.get("expected", 1)
                    # attribute = raw measurement; Column = did appraiser agree with standard?
                    rows.append({
                        "Column":     expected,
                        "Samples":    s,
                        "Appraisers": app_insp_map.get(appraiser, appraiser),
                        "trial":      t,
                        "attribute":  attr_lookup.get((t, s), expected),
                    })

        return pd.DataFrame(rows)

    def _build_summary(
        self,
        sample_map: Dict[int, Dict],
        df_sheet1: pd.DataFrame,
        appraisers: List[str]
    ) -> List[List]:
        """
        Build Summary sheet rows: Item, Error Code, Standard, A-1..C-3 (9 columns).
        For each sample: expected value + actual results per trial.
        """
        rows = []
        # Determine trial ordering: A-1..A-3, B-1..B-3, C-1..C-3
        trial_order = [f"{a}-{t}" for a in appraisers for t in range(1, 4)]

        for sample_num in range(1, 11):
            info = sample_map.get(sample_num, {})
            item      = info.get("item_name", f"Sample {sample_num}")
            err_code  = info.get("error_code", "PASS")
            expected  = info.get("expected", 1)

            row = [item, err_code, expected]
            # For each trial slot, look up actual from sheet1
            for trial in trial_order:
                app = trial.split("-")[0]
                trial_num = int(trial.split("-")[1])
                sub = df_sheet1[
                    (df_sheet1["Samples"] == sample_num) &
                    (df_sheet1["Appraisers"] == app) &
                    (df_sheet1["trial"] == trial_num)
                ]
                if len(sub) > 0:
                    val = int(sub["attribute"].iloc[0])
                else:
                    val = expected
                row.append(val)
            rows.append(row)
        return rows

    def generate(
        self,
        df_jason: pd.DataFrame,
        sample_map: Dict[int, Dict],
        part_number: str = None,
        instrument: str = None,
        instrument_no: str = None,
        department: str = None,
        reported_by: str = None,
        project_name: str = None,
        inspector_numbers: Optional[List[str]] = None,
        fixture_no: str = None,
        date_inspected: Optional[str] = None,
        faults_map: Optional[Dict[int, tuple]] = None,
        n_trials: int = 3,
        template_path: Optional[str] = str(Path(__file__).parent / "AR_R_Template.xlsx"),
        chart_paths: Optional[dict] = None,
        layout: Optional[list] = None,
    ) -> str:
        """
        Generate complete AR&R report.

        Args:
            df_jason:       Parsed DataFrame from parse_json_folder
            sample_map:     {sample_num: {"item_name": str, "error_code": str, "expected": 0 or 1}}
                            e.g. {1: {"item_name": "Known bad DUT 1", "error_code": "CONTROL_MT_MODE_STATE", "expected": 0}}
            faults_map:     {row_idx: (error_code, description)} for Faults simulation sheet
            template_path:  Path to Form-004091 template Excel to base report on
            layout:         Override ARR_LAYOUT (from utils.config) for chart placement.
                            List of (filename, col, row, w_px, h_px, colOff_px, rowOff_px).
        Returns:
            Path to saved .xlsx file
        """
        # Resolve defaults from config when caller didn't supply a value
        _d = ARR_FORM_DEFAULTS
        part_number    = part_number    if part_number    is not None else _d["part_number"]
        description    = _d["description"]
        instrument     = instrument     if instrument     is not None else _d["instrument"]
        instrument_no  = instrument_no  if instrument_no  is not None else _d["instrument_no"]
        department     = department     if department     is not None else _d["department"]
        reported_by    = reported_by    if reported_by    is not None else _d["reported_by"]
        fixture_no     = fixture_no     if fixture_no     is not None else _d["fixture_no"]
        project_name   = project_name   if project_name   is not None else _d["project_name"]
        if not inspector_numbers:
            inspector_numbers = _d["inspector_numbers"] or []

        wb = self._ensure_template()

        date_str = date_inspected or datetime.now().strftime("%b %d,%Y")

        # ── Identify appraisers ────────────────────
        appraisers = sorted(df_jason["appraiser"].dropna().unique().tolist())
        n_app = len(appraisers)
        if n_app == 0:
            raise ValueError("No appraisers found in Jason logs")

        # Auto-fill inspector_numbers from JSON appraiser data if not provided
        if not inspector_numbers:
            inspector_numbers = appraisers
        inspector_numbers = [str(n) for n in inspector_numbers]

        # Assign appraiser letters: A, B, C, ...
        app_letters = {}
        for i, app in enumerate(appraisers):
            app_letters[app] = chr(ord("A") + i)

        # ── Build Sheet1 ────────────────────────────
        logger.info("generate: inspector_numbers=%s, n_trials=%s", inspector_numbers, n_trials)
        df_s1 = ARRReportGenerator._build_sheet1(df_jason, sample_map, n_trials, inspector_numbers)

        # ── Compute attribute agreement ────────────
        metrics = compute_attribute_agreement(df_s1, None)

        # ── Fill Cover Page ────────────────────────
        ws = wb["Cover Page"] if "Cover Page" in wb.sheetnames else wb.active
        # V258: Inspector ID values are no longer written into
        # the Cover Page rows 5/6/7 column 3 (per user request).
        # The cells are left blank so the header block still
        # has its layout but the inspector numbers don't show.
        # if inspector_numbers:
        #     insp = inspector_numbers[:3]
        #     for i, num in enumerate(insp):
        #         r = 5 + i
        #         if r <= 7:
        #             ws.cell(r, 3).value = num
        #             ws.cell(r, 3).border = _thin()
        #             ws.cell(r, 3).font   = Font(name="Calibri", size=9)

        # ── Fill AR&R Form ─────────────────────────
        if "AR&R Form" in wb.sheetnames:
            ws = wb["AR&R Form"]
            # Metadata — value only (border/font/fill preserved from template)
            ws.cell(2, 3).value = part_number
            ws.cell(2, 9).value = description
            ws.cell(3, 3).value = instrument
            ws.cell(3, 9).value = instrument_no
            ws.cell(4, 3).value = department
            ws.cell(4, 9).value = reported_by
            ws.cell(5, 3).value = int(inspector_numbers[0]) if inspector_numbers else "—"
            ws.cell(5, 9).value = fixture_no
            ws.cell(6, 3).value = int(inspector_numbers[1]) if len(inspector_numbers) > 1 else "—"
            ws.cell(6, 9).value = project_name
            ws.cell(7, 3).value = int(inspector_numbers[2]) if len(inspector_numbers) > 2 else "—"
            ws.cell(7, 9).value = date_str

            # Data rows 11–20: read actual measurements from df_jason directly
            # (df_jason has all 90 trial records; df_s1 has only 1 per appraiser×sample)
            # appraiser letters: "A", "B", "C" (match df_jason["appraiser"] values)
            trial_slots = [f"{chr(ord('A')+i)}-{t}" for i in range(n_app) for t in range(1, 4)]
            for sample_num in range(1, 11):
                r = 10 + sample_num
                info = sample_map.get(sample_num, {})
                expected = info.get("expected", 1)
                for j, trial in enumerate(trial_slots):
                    app_letter = trial.split("-")[0]   # "A", "B", "C"
                    trial_num  = int(trial.split("-")[1])  # 1, 2, or 3
                    # Look up from df_jason (all 90 trial records)
                    sub = df_jason[
                        (df_jason["appraiser"] == app_letter) &
                        (df_jason["sample_num"] == sample_num) &
                        (df_jason["trial"] == trial_num)
                    ]
                    if len(sub) > 0:
                        val = int(sub["attribute"].iloc[0])
                    else:
                        val = expected
                    # Use _val() to preserve template styles (font/fill/border)
                    _val(ws.cell(r, j + 2), val, fmt="0")
                # Basis column: preserve template right-medium border
                _val(ws.cell(r, 11), expected, fmt="0")

        # ── Fill Sheet1 ────────────────────────────
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            for i, (_, row) in enumerate(df_s1.iterrows(), start=2):
                ws.cell(i, 1).value = int(row["Column"])
                ws.cell(i, 2).value = int(row["Samples"])
                ws.cell(i, 3).value = str(row["Appraisers"])
                ws.cell(i, 4).value = int(row["attribute"])
                for c in range(1, 5):
                    ws.cell(i, c).border = _thin()
                    ws.cell(i, c).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(i, c).font = Font(name="Arial", size=10)
                # Colour: green if correct, red if wrong
                if row["Column"] == row["attribute"]:
                    ws.cell(i, 4).fill = _fill(CLR_GRN)
                else:
                    ws.cell(i, 4).fill = _fill(CLR_RED)

        # ── Fill Summary ────────────────────────────
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            summary_rows = self._build_summary(sample_map, df_s1, appraisers)
            trial_slots = [f"{chr(ord('A')+i)}-{t}" for i in range(n_app) for t in range(1, 4)]
            # Row 1 header
            ws.cell(1, 1).value = "Item"
            ws.cell(1, 2).value = "Error Code"
            ws.cell(1, 3).value = "Standard"
            for j, t in enumerate(trial_slots, 4):
                ws.cell(1, j).value = t
                ws.cell(1, j).fill  = _fill(CLR_SUBHDR_BG)
                ws.cell(1, j).font  = Font(name="Calibri", bold=True, size=9, color=CLR_SUBHDR_FG)
                ws.cell(1, j).border = _thin()
                ws.cell(1, j).alignment = Alignment(horizontal="center", vertical="center")

            for r, srow in enumerate(summary_rows, 3):
                ws.cell(r, 1).value = srow[0]  # Item
                ws.cell(r, 2).value = srow[1]  # Error Code
                ws.cell(r, 3).value = srow[2]  # Standard
                for j, val in enumerate(srow[3:], 4):
                    ws.cell(r, j).value = val
                    ws.cell(r, j).border = _thin()
                    ws.cell(r, j).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(r, j).font = Font(name="Calibri", size=9)
                # Colour: green if standard=actual, red otherwise
                std = srow[2]
                for j, val in enumerate(srow[3:], 4):
                    bg = CLR_GRN if val == std else CLR_RED
                    ws.cell(r, j).fill = _fill(bg)

        # ── Fill Faults simulation ─────────────────
        if faults_map and "Faults simulation" in wb.sheetnames:
            ws = wb["Faults simulation"]
            for row_idx, (code, desc) in faults_map.items():
                ws.cell(row_idx + 2, 1).value = row_idx
                ws.cell(row_idx + 2, 2).value = code
                ws.cell(row_idx + 2, 3).value = desc
                for c in range(1, 4):
                    ws.cell(row_idx + 2, c).border = _thin()

        # ── Fill AR&R Report ───────────────────────
        if "AR&R Report" in wb.sheetnames:
            ws = wb["AR&R Report"]
            overall_eff = metrics.get("overall_eff", 0)
            by_app      = metrics.get("vs_standard", {})
            between_eff = metrics.get("between_appraiser", 0)

            # Conclusion table (rows 16–19, cols N=A, O=B, P=C)
            conclusion_vals = [
                (16, list(metrics.get("within_appraiser", {}).values())),  # within
                (17, list(by_app.values())),                                # vs standard
                (18, [between_eff]),                                        # between
                (19, [overall_eff]),                                        # all vs standard
            ]
            for row_num, eff_list in conclusion_vals:
                for i, eff in enumerate(eff_list[:3]):
                    col = 13 + i   # M=A(13), N=B(14), O=C(15)
                    pct = round(eff, 1) if eff else 0.0
                    # Template already has correct style (fill, font, border) for PASS/FAIL
                    # Only write value so template style is preserved
                    label = "PASS" if pct >= 90 else "FAIL"
                    ws.cell(row_num, col).value = label


            # ── Embed Minitab AR&R charts into AR&R Report sheet ───
            if chart_paths and "AR&R Report" in wb.sheetnames:
                ws_rep = wb["AR&R Report"]
                try:
                    from openpyxl.drawing.image import Image as _Img
                    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker

                    # Clear all existing template images first
                    removed = 0
                    if hasattr(ws_rep, '_images') and ws_rep._images:
                        removed = len(ws_rep._images)
                        ws_rep._images.clear()
                        logger.info("Removed %d template images from AR&R Report", removed)

                    # Layout: (filename, col, row, width_px, height_px, colOff_px, rowOff_px)
                    # Columns are 1-indexed: B=2, G=7, L=12
                    # Position defaults are loaded from utils.config.ARR_LAYOUT;
                    # can be overridden per-run by passing `layout` to generate().
                    _layout = chart_paths.get("_layout") or ARR_LAYOUT

                    img_dir = chart_paths.get("_img_dir", "")
                    for img_name, col, row, w, h, colOff_px, rowOff_px in _layout:
                        img_path = str(Path(img_dir) / img_name) if img_dir else img_name
                        if not Path(img_path).exists():
                            logger.warning("Chart image not found, skipping: %s", img_path)
                            continue

                        img = _Img(img_path)
                        img.width  = w
                        img.height = h
                        # OneCellAnchor requires ext (size in EMU). 1 px = 9525 EMU at 96 DPI.
                        from openpyxl.drawing.xdr import XDRPositiveSize2D
                        anchor = OneCellAnchor(
                            _from=AnchorMarker(
                                col=col - 1, colOff=int(colOff_px * 9525),
                                row=row - 1, rowOff=int(rowOff_px * 9525),
                            ),
                            ext=XDRPositiveSize2D(cx=int(w * 9525), cy=int(h * 9525)),
                        )
                        img.anchor = anchor
                        ws_rep.add_image(img)
                        logger.info("AR&R chart '%s' embedded at R%dC%d (%dx%d)",
                                     img_name, row, col, w, h)
                except Exception as e:
                    logger.warning("Failed to embed AR&R charts: %s", e)

        # ── Save ────────────────────────────────────
        date_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"AR_R_Report_{date_ts}.xlsx"
        out_path = self.output_dir / filename
        wb.save(out_path)
        logger.info(f"AR&R report saved: {out_path}")

        # ── Cleanup extracted images ──────────────────
        if chart_paths and isinstance(chart_paths, dict):
            try:
                extracted = chart_paths.get("_extracted", [])
                img_dir  = chart_paths.get("_img_dir", "")
                for p in extracted:
                    try:
                        Path(p).unlink()
                        logger.info("Deleted image: %s", p)
                    except Exception:
                        pass
                if img_dir and Path(img_dir).exists():
                    try:
                        Path(img_dir).rmdir()
                        logger.info("Deleted img_dir: %s", img_dir)
                    except Exception:
                        pass
            except Exception as e:
                logger.warning("Image cleanup failed: %s", e)

        return str(out_path)


# ──────────────────────────────────────────────
CLR_SECTION_BG = "D6E4F0"
