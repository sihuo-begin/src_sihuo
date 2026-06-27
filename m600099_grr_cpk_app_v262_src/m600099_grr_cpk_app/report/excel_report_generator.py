# ──────────────────────────────────────────────
#  Excel Report Generator
#  Produces a FORM-004090 GRR Excel report
# ──────────────────────────────────────────────
import logging
from datetime import datetime
import os
import re
import json
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from utils.config import (
    LED_SPECS,
    PNUM_MAPPING,
    _norm_led,
    GRR_CHART_LAYOUT,
    GRR_SHEET_NAMING,
    GRR_TEMPLATE_PATH,
    GRR_FORM_DEFAULTS,
    GRR_LAYOUT,
    OUTPUT_DIR,
)
from core.json_parser import JsonParser

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────
CLR_HEADER_BG = "1F4E79"  # dark blue
CLR_HEADER_FG = "FFFFFF"
CLR_SUBHDR_BG = "2E75B6"
CLR_SUBHDR_FG = "FFFFFF"
CLR_SECTION_BG = "D6E4F0"
CLR_PASS_BG = "E2EFDA"  # light green
CLR_FAIL_BG = "FCE4D6"  # light red
CLR_WARN_BG = "FFF2CC"  # light yellow
CLR_ALT_ROW = "EBF3FB"  # alternating row
CLR_NOTE_BG = "FFFACD"
CLR_GRR_GOOD = "C6EFCE"
CLR_GRR_ACC = "FFEB9C"
CLR_GRR_BAD = "FFC7CE"


# ── Helper styles ───────────────────────────────
def _hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _bold_border() -> Border:
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(cell, text: str, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG, size=10, bold=True, wrap=False, align="center"):
    cell.value = text
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = _hdr_fill(bg)
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def _apply_data(cell, value, fmt=None, bold=False, align="center", bg=None, fg="000000"):
    cell.value = value
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    if bg:
        cell.fill = _hdr_fill(bg)
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


def _grr_result_color(pct_grr: float) -> tuple:
    if pct_grr <= 10:
        return CLR_GRR_GOOD, "000000", "Good"
    elif pct_grr <= 30:
        return CLR_GRR_ACC, "9C6500", "Acceptable"
    else:
        return CLR_GRR_BAD, "9C0006", "Bad"


def _result_pass_fail(pct_grr: float) -> str:
    if pct_grr <= 30:
        return "PASS"
    return "FAIL"


# ──────────────────────────────────────────────
#  Main Report Generator
# ──────────────────────────────────────────────
class ExcelReportGenerator(JsonParser):
    """
    Generates a FORM-004090 GRR Excel report matching the company template.

    Sheet structure:
      1. Cover Page   – approvals + revision history
      2. GRR Form     – raw measurement data (Sample, Inspector, PNUM-4024~PNUM-4056)
      3. Summary       – %GRR / %PT / NDC results for all parameters
      4..N  One sheet per item, named by 4-digit Pnum value (e.g. "4024"),
             containing GRR metrics + embedded Minitab GageRR chart.
    """

    def __init__(self, output_dir: str = "output", template_path: Optional[str] = GRR_TEMPLATE_PATH):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.template_path = template_path
        self._used_template = False  # set by _load_or_create_template

    def _load_or_create_template(self) -> openpyxl.Workbook:
        """Load GRR template workbook if it exists, else create a blank one.
        Preserves formatting and the first 4 sheets (Cover Page, GRR Form,
        Summary, Summary) when the template is used.
        """
        if self.template_path and Path(self.template_path).exists():
            try:
                wb = openpyxl.load_workbook(self.template_path)
                self._used_template = True
                logger.info("GRR template loaded: %s", self.template_path)
                return wb
            except Exception as e:
                logger.warning("Failed to load GRR template %s: %s. Falling back to blank.", self.template_path, e)
        self._used_template = False
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        logger.info("GRR template not found at %s; using blank workbook.", self.template_path)
        return wb

    def generate(
        self,
        results: Dict,
        df: pd.DataFrame,
        inline_charts: bool = False,
        minitab_path: Optional[str] = None,
        product_name: str = "Hermes Min Charger",
        station: str = "MT7",
        project_name: str = "Beta",
        reported_by: str = "Min Duan",
        inspectors: Optional[List[str]] = None,
    ) -> str:
        """
        Generate the full Excel GRR report.

        Args:
            results:       {item_name: {"grr": GRRResult, "cpk": CPKResult}}
            df:            Full GRR DataFrame (Sample, Inspector, PNUM cols)
            inline_charts: Whether to embed Minitab charts
            minitab_path:  Path to Minitab Mtb.exe
            product_name:  Product name for report header
            station:       Station ID (e.g. MT7)
            project_name:  Project name
            reported_by:   Who generated the report
            inspectors:    List of inspector numbers (auto-detected if None)

        Returns:
            Path to the saved .xlsx file
        """
        # Try template first; fall back to building from scratch.
        wb = self._load_or_create_template()
        print("simon result, df is {} {} ".format(results, df))
        # ── Identify per-item sheets in the template. ──
        # Two template styles are supported:
        #   A. Full template — one pre-built sheet per item (PNUM-XXXX).
        #      We update them in place (preserves formatting perfectly).
        #   B. Minimal template — one generic "PNUM-TEMPLATE" sheet that we
        #      copy + rename for each item. Used when the user prefers a
        #      smaller template file.
        template_pnum_sheets = {}  # {pnum_str: ws}     for style A
        template_pnum_source = None  # ws to copy from     for style B
        if self._used_template:
            for ws_name in list(wb.sheetnames):
                m = re.match(r"^PNUM-(\d{4})$", ws_name) or re.match(r"^(\d{4})$", ws_name)
                if m:
                    template_pnum_sheets[m.group(1)] = wb[ws_name]
            # Look for a generic source sheet to copy
            for candidate in ("PNUM-TEMPLATE", "PNUM_TEMPLATE", "PNUM-T", "PNUM-Template"):
                if candidate in wb.sheetnames:
                    template_pnum_source = wb[candidate]
                    break
            if template_pnum_source is None and not template_pnum_sheets:
                # Fall back: any per-item-shaped sheet works as source
                for ws_name in list(wb.sheetnames):
                    if ws_name in ("Cover Page", "GRR From", "GRR Form", "Summary ", "Summary"):
                        continue
                    if re.match(r"^PNUM-\d{4}$", ws_name) or re.match(r"^\d{4}$", ws_name):
                        template_pnum_source = wb[ws_name]
                        break

        # ── Gather metadata ──
        grr_items = [k for k, v in results.items() if "grr" in v and v["grr"]]
        # Prefer inspector_numbers from GRR_FORM_DEFAULTS (the canonical
        # work-order IDs). Fall back to the GRR Form template's C5/C6/C7
        # (these encode the Minitab Set C2 cycling order), then to the
        # `inspectors` arg, and finally to auto-detected values from df.
        cfg_inspectors = list(GRR_FORM_DEFAULTS.get("inspector_numbers") or [])
        template_inspectors = []
        if not cfg_inspectors and self._used_template:
            for cand in ("GRR Form", "GRR From"):
                if cand in wb.sheetnames:
                    ws_meta = wb[cand]
                    for r in (5, 6, 7):
                        v = ws_meta.cell(r, 3).value
                        if v is not None and str(v).strip():
                            template_inspectors.append(str(v).strip())
                    break
        if cfg_inspectors:
            inspector_numbers = [str(x) for x in cfg_inspectors[:3]]
        elif template_inspectors:
            inspector_numbers = template_inspectors[:3]
        elif inspectors:
            inspector_numbers = [str(i) for i in inspectors]
        else:
            # Auto-detect inspectors from df
            insp_col = next((c for c in df.columns if c.lower() in ("inspector", "appraiser")), None)
            if insp_col:
                inspector_numbers = sorted(df[insp_col].dropna().unique().tolist(), key=str)
            else:
                inspector_numbers = []

        date_str = datetime.now().strftime("%B %d, %Y")
        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")

        # ── 1. Cover Page (only create if no template) ──
        if not self._used_template or "Cover Page" not in wb.sheetnames:
            self._make_cover_page(wb, product_name, station, project_name, reported_by, inspector_numbers, date_str)

        # ── 2. GRR Form / GRR From sheet ──
        # When the template is loaded, KEEP its existing sheet (either
        # "GRR From" or "GRR Form") and update the values in place.
        # Only create a brand-new sheet when the template has no GRR
        # Form at all.
        existing_grr_form = None
        if self._used_template:
            for candidate in ("GRR Form", "GRR From"):
                if candidate in wb.sheetnames:
                    existing_grr_form = wb[candidate]
                    break
        if existing_grr_form is not None:
            self._update_grr_form_page(existing_grr_form, df, inspector_numbers, date_str, selected_led_cols=grr_items)
        else:
            self._make_grr_data_page(wb, df, inspector_numbers)

        # ── 3. Summary sheets ──
        for summary_name in ("Summary ", "Summary"):
            if self._used_template and summary_name in wb.sheetnames:
                self._update_summary_page(wb[summary_name], results, grr_items, product_name, station, project_name)
            elif not self._used_template and summary_name not in wb.sheetnames:
                if summary_name == "Summary ":
                    self._make_summary_page(wb, results, grr_items)

        # ── 4a. Remove unselected PNUM sheets from template output ──
        # If the template ships with 9 pre-made PNUM-XXXX sheets but the
        # user only ran a few items, drop the unused ones so the report
        # only contains sheets for the items actually processed.
        # if self._used_template:
        #     selected_pnum_strs = set()
        #     for it in grr_items:
        #         pnum_raw = self.LED_SPECS.get(it, {}).get("pnum", it[-4:])
        #         pnum_s = str(pnum_raw) if not isinstance(pnum_raw, str) else pnum_raw
        #         selected_pnum_strs.add(pnum_s)
        #     for pnum_s, ws in list(template_pnum_sheets.items()):
        #         if pnum_s not in selected_pnum_strs:
        #             try:
        #                 wb.remove(ws)
        #                 del template_pnum_sheets[pnum_s]
        #                 logger.info(f"Removed unselected PNUM sheet: {ws.title}")
        #             except Exception as ex:
        #                 logger.warning(f"Could not remove {ws.title}: {ex}")

        # ── 4. Per-parameter sheets ──
        for item in grr_items:
            grr_res = results[item]["grr"]
            spec = self.LED_SPECS.get(item, {})
            chart_paths_for_item = grr_res.chart_paths if hasattr(grr_res, "chart_paths") else {}

            pnum_raw = self.LED_SPECS.get(item, {}).get("pnum", item[-4:])
            pnum_str = str(pnum_raw) if not isinstance(pnum_raw, str) else pnum_raw

            # If template already has a sheet for this PNUM, update values
            # + replace images. Otherwise, if there's a generic source sheet
            # (PNUM-TEMPLATE), copy it and rename. Otherwise create a new
            # sheet.
            existing_ws = template_pnum_sheets.get(pnum_str)
            if existing_ws is not None and self._used_template:
                self._update_pnum_page(existing_ws, item, grr_res, spec, chart_paths_for_item)
            elif template_pnum_source is not None and self._used_template:
                # Copy the generic template sheet, rename, then update
                new_ws = wb.copy_worksheet(template_pnum_source)
                new_ws.title = f"PNUM-{pnum_str}"[:31]
                self._update_pnum_page(new_ws, item, grr_res, spec, chart_paths_for_item)
            else:
                self._make_pnum_page(wb, item, grr_res, spec, chart_paths_for_item)

        # ── Save ──
        filename = f"Beta-GRR-Charger_{station}-HVTE-M600099-{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        out_path = self.output_dir / filename
        wb.save(out_path)
        logger.info(f"GRR Excel report saved: {out_path}")
        return str(out_path)

    # ── Cover Page ────────────────────────────────

    def _make_cover_page(self, wb, product, station, project, reported_by, inspectors, date_str):
        ws = wb.create_sheet("Cover Page")
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20

        # Row 1-2: blank (spacer)
        # Row 3-5: Approvals header
        for r in range(3, 11):
            ws.row_dimensions[r].height = 18

        _apply_header(ws.cell(3, 1), "APPROVALS", bg=CLR_HEADER_BG, size=11, wrap=True)
        _apply_header(ws.cell(3, 2), "DEPT", bg=CLR_HEADER_BG)
        _apply_header(ws.cell(3, 3), "NAME", bg=CLR_HEADER_BG)
        _apply_header(ws.cell(3, 4), "TITLE", bg=CLR_HEADER_BG)
        _apply_header(ws.cell(3, 5), "DATE", bg=CLR_HEADER_BG)

        approval_data = [
            ("QA", "Kerwin Wang", "Asst. Manager", "In EQMS"),
        ]
        for i, row in enumerate(approval_data, start=4):
            for j, val in enumerate(row, start=2):
                _apply_data(ws.cell(i, j), val, align="center")

        # Approval rows (blank)
        for r in range(5, 11):
            for c in range(1, 6):
                ws.cell(r, c).border = _thin_border()

        # ── Title block (columns F-G merged) ──
        ws.merge_cells("F3:G3")
        _apply_header(ws.cell(3, 6), "TEMPLATE TITLE", bg=CLR_HEADER_BG, size=11, wrap=True)

        ws.merge_cells("F4:G4")
        _apply_header(ws.cell(4, 6), "Gage R and R Form", bg=CLR_SUBHDR_BG, size=12, bold=True)

        ws.merge_cells("F5:G5")
        _apply_header(ws.cell(5, 6), "Gage R and R Form", bg=CLR_SUBHDR_BG, size=12, bold=True)

        ws.merge_cells("F6:G6")
        _apply_header(ws.cell(6, 6), "TEMPLATE NUMBER FORM-004090", bg=CLR_SUBHDR_BG, size=10)

        ws.merge_cells("F7:G7")
        _apply_header(ws.cell(7, 6), "REVISION B", bg=CLR_SUBHDR_BG, size=10)

        # ── Revision history ──
        ws.merge_cells("A10:G10")
        _apply_header(ws.cell(10, 1), "REVISION HISTORY", bg=CLR_HEADER_BG, size=10, wrap=True)

        rev_headers = ["REVISION", "DESCRIPTION OF CHANGE", "ORIGINATOR", "RELEASE DATE"]
        for j, h in enumerate(rev_headers, start=1):
            _apply_header(ws.cell(11, j), h, bg=CLR_SUBHDR_BG)
        for j in range(5, 8):
            ws.cell(11, j).border = _thin_border()

        rev_data = [
            ("A", "Transfer from QDMS to EQMS", "Li Rao", "In EQMS"),
            ("B", "Update header and footer for the template", "Li Rao", "In EQMS"),
        ]
        for i, row in enumerate(rev_data, start=12):
            for j, val in enumerate(row, start=1):
                _apply_data(ws.cell(i, j), val, align="left" if j == 2 else "center")
            for j in range(5, 8):
                ws.cell(i, j).border = _thin_border()

        # Note row
        ws.merge_cells("A14:G14")
        note_cell = ws.cell(
            14,
            1,
            value="Note: The cover page is only used for electronic file traceability of change history, "
            "no need print out the cover page when do hardcopy record",
        )
        note_cell.font = Font(name="Calibri", italic=True, size=8, color="595959")
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[14].height = 30

        # ── Freeze pane ──
        ws.freeze_panes = "A3"

    # ── GRR Raw Data Page ────────────────────────

    def _make_grr_data_page(self, wb, df: pd.DataFrame, inspector_numbers: List[str]):
        ws = wb.create_sheet("GRR Form")

        # Column widths
        ws.column_dimensions["A"].width = 8  # Sample
        ws.column_dimensions["B"].width = 12  # Inspector
        for col_idx in range(3, 12):  # PNUM columns
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # ── Header block (rows 1-11) ──
        ws.merge_cells("A1:K1")
        ws.merge_cells("A2:K2")

        ws.merge_cells("A3:C3")
        _apply_header(ws.cell(3, 1), "Part Number :", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        ws.cell(3, 4).value = GRR_FORM_DEFAULTS["part_number"]
        _apply_data(ws.cell(3, 4), GRR_FORM_DEFAULTS["part_number"], align="center")
        ws.merge_cells("D3:K3")

        ws.merge_cells("A4:C4")
        _apply_header(ws.cell(4, 1), "Instrument name:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        ws.merge_cells("D4:G4")
        _apply_data(ws.cell(4, 4), GRR_FORM_DEFAULTS["instrument"], align="center")

        ws.merge_cells("H4:I4")
        _apply_header(ws.cell(4, 8), "Instrument No:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(4, 10), GRR_FORM_DEFAULTS["instrument_no"], align="center")

        ws.merge_cells("A5:C5")
        _apply_header(ws.cell(5, 1), "Department Name:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(5, 4), GRR_FORM_DEFAULTS["department"], align="center")

        ws.merge_cells("H5:I5")
        _apply_header(ws.cell(5, 8), "Reported by:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(5, 10), reported_by_val := GRR_FORM_DEFAULTS["reported_by"], align="center")

        # Inspector Number: 3 rows, one per inspector.
        # Each row shows the inspector's ID once (no repeat across columns).
        for row_idx, insp in enumerate(inspector_numbers[:3], start=6):
            ws.merge_cells(f"A{row_idx}:C{row_idx}")
            _apply_header(ws.cell(row_idx, 1), "Inspector Number:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
            _apply_data(ws.cell(row_idx, 4), str(insp), align="center")

        ws.merge_cells("H6:I6")
        _apply_header(ws.cell(6, 8), "Measurement Unit:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(6, 10), GRR_FORM_DEFAULTS["measurement_unit"], align="center")

        ws.merge_cells("H7:I7")
        _apply_header(ws.cell(7, 8), "Project Name:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(7, 10), GRR_FORM_DEFAULTS["project_name"], align="center")

        ws.merge_cells("H8:I8")
        _apply_header(ws.cell(8, 8), "Date Inspected:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(8, 10), datetime.now().strftime("%B %d, %Y"), align="center")

        # ── Column headers (row 9) ──
        col_headers = (
            ["Parameter"] * 2
            + ["Sample", "Inspector"]
            + [f"PNUM-{p}" for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]
        )
        for j, h in enumerate(col_headers, start=1):
            _apply_header(ws.cell(9, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=10)
        ws.merge_cells("A9:B9")
        _apply_header(ws.cell(9, 1), "Parameter", bg=CLR_SUBHDR_BG)
        ws.cell(9, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("C9:D9")
        _apply_header(ws.cell(9, 3), "Parameter", bg=CLR_SUBHDR_BG)

        # Row 10: column sub-headers
        _apply_header(ws.cell(10, 1), "Sample", bg=CLR_SUBHDR_BG)
        _apply_header(ws.cell(10, 2), "Inspector", bg=CLR_SUBHDR_BG)
        for j in range(3, 12):
            ws.cell(10, j).value = None
            ws.cell(10, j).border = _thin_border()

        # ── Data rows ──
        # Determine which column holds inspector numbers
        insp_col = None
        sample_col = None
        pnum_col_map = {}
        for col in df.columns:
            cl = col.strip().lower()
            if cl in ("inspector", "appraiser"):
                insp_col = col
            elif cl == "sample":
                sample_col = col
            elif col in PNUM_MAPPING.values() or any(p in col for p in ["PNUM-", "INTENSITY"]):
                # Match PNUM column
                for pnum, led in PNUM_MAPPING.items():
                    if pnum in col or led in col:
                        pnum_col_map[pnum] = col
                        break

        # If df is already the GRR From style (PNUM-4024 style columns)
        pnum_in_df = [c for c in df.columns if c.startswith("PNUM-")]

        start_row = 11
        row_idx = start_row

        if pnum_in_df:
            # GRR template format: Sample, Inspector, PNUM-4024, ...
            pnums = ["PNUM-" + str(p) for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]
            df_out = df[[c for c in pnums if c in df.columns]]
            insp_vals = df["Inspector"].tolist() if "Inspector" in df.columns else []
            sample_vals = df["Sample"].tolist() if "Sample" in df.columns else []

            for i, (idx, row) in enumerate(df_out.iterrows()):
                bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
                sn_val = sample_vals[i] if i < len(sample_vals) else ""
                insp_val = insp_vals[i] if i < len(insp_vals) else ""

                _apply_data(ws.cell(row_idx, 1), sn_val, bg=bg, align="center")
                _apply_data(ws.cell(row_idx, 2), insp_val, bg=bg, align="center")
                for j, pnum in enumerate(pnums):
                    val = row.get(pnum, None)
                    if pd.notna(val):
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            val = None
                    _apply_data(ws.cell(row_idx, 3 + j), val, bg=bg, align="center", fmt="0")
                row_idx += 1

        else:
            # Intermediate format: need to pivot
            # df has: sn, appraiser, LED_XXX_INTENSITY columns
            led_to_pnum = {v: k for k, v in PNUM_MAPPING.items()}

            # Determine unique parts and operators
            sn_col = (
                [c for c in df.columns if c.lower() in ("sn", "part_num", "sample")][0]
                if any(c.lower() in ("sn", "part_num", "sample") for c in df.columns)
                else None
            )
            op_col = (
                [c for c in df.columns if c.lower() in ("appraiser", "inspector", "operator")][0]
                if any(c.lower() in ("appraiser", "inspector", "operator") for c in df.columns)
                else None
            )

            if sn_col and op_col:
                df = df.copy()
                if sn_col == "part_num":
                    pass  # already numbered
                elif sn_col == "sn":
                    unique_sns = sorted(df[sn_col].dropna().unique(), key=lambda x: str(x))
                    sn_map = {s: i + 1 for i, s in enumerate(unique_sns)}
                    df["part_num"] = df[sn_col].map(sn_map)

                unique_ops = sorted(df[op_col].dropna().unique(), key=str)
                op_map = {o: str(o) for o in unique_ops}
                df["op_str"] = df[op_col].map(op_map)

                # Build GRR table rows
                led_cols = [c for c in df.columns if "INTENSITY" in c]
                for pnum in ["PNUM-" + str(p) for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]:
                    pass  # will fill below

                # Write header row for each pnum column
                pnum_headers = ["PNUM-" + str(p) for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]
                for j, ph in enumerate(pnum_headers):
                    _apply_header(ws.cell(10, 3 + j), ph, bg=CLR_SECTION_BG, fg="1F4E79", size=9)

                row_idx = start_row
                for part in range(1, 11):
                    for op in unique_ops:
                        # Pull every measurement for this (part, op); if
                        # multiple trials exist, include all of them in
                        # separate rows so Minitab can group them.
                        sub = df[(df["part_num"] == part) & (df["op_str"] == str(op))]
                        if sub.empty:
                            continue
                        # If the data has no "trial" column, treat each row
                        # as one trial and emit up to 3 of them.
                        for trial_idx, (_, r) in enumerate(sub.head(3).iterrows(), start=1):
                            bg = CLR_ALT_ROW if (part + trial_idx) % 2 == 0 else "FFFFFF"
                            _apply_data(ws.cell(row_idx, 1), part, bg=bg)
                            _apply_data(ws.cell(row_idx, 2), str(op), bg=bg)
                            for j, led in enumerate(led_cols):
                                val = r.get(led, None)
                                if pd.notna(val):
                                    try:
                                        val = float(val)
                                    except (ValueError, TypeError):
                                        val = None
                                pnum_col = PNUM_MAPPING_REV.get(led, "")
                                col_idx = pnum_headers.index(pnum_col) + 3 if pnum_col in pnum_headers else -1
                                if col_idx >= 0:
                                    _apply_data(ws.cell(row_idx, col_idx), val, bg=bg, fmt="0")
                            row_idx += 1

        # ── Footer ──
        footer_row = row_idx + 1
        ws.merge_cells(f"A{footer_row}:K{footer_row}")
        _apply_data(ws.cell(footer_row, 1), "This Template Link to #WI-002388", align="left", bold=True, fg="595959")

        ws.freeze_panes = "A11"

    def _update_grr_form_page(self, ws, df, inspector_numbers, date_str, selected_led_cols=None):
        """Update values in the EXISTING template 'GRR Form' or 'GRR From' sheet.
        Layout (matches the template):
          Row 2:  C2=Part Number, I2=Description
          Row 3:  C3=Instrument, I3=Instrument No
          Row 4:  C4=Department, I4=Reported by
          Row 5:  C5=Inspector 1, I5=Measurement Unit
          Row 6:  C6=Inspector 2, I6=Project Name
          Row 7:  C7=Inspector 3, I7=Date Inspected
          Row 10: headers (Sample | Inspector | PNUM-4024 … PNUM-4056)
          Row 11+: data rows (up to 90 rows for 3 inspectors × 30 trials)
        """
        # Header metadata
        ws.cell(2, 3).value = GRR_FORM_DEFAULTS["part_number"]
        ws.cell(2, 9).value = GRR_FORM_DEFAULTS.get("description", "Beta Charger")
        ws.cell(3, 3).value = GRR_FORM_DEFAULTS["instrument"]
        ws.cell(3, 9).value = GRR_FORM_DEFAULTS["instrument_no"]
        ws.cell(4, 3).value = GRR_FORM_DEFAULTS["department"]
        ws.cell(4, 9).value = GRR_FORM_DEFAULTS["reported_by"]
        for i, insp in enumerate(inspector_numbers[:3]):
            ws.cell(5 + i, 3).value = str(insp)
        ws.cell(5, 9).value = GRR_FORM_DEFAULTS["measurement_unit"]
        ws.cell(6, 9).value = GRR_FORM_DEFAULTS["project_name"]
        ws.cell(7, 9).value = date_str
        self.pnum_item_map = {}
        self.LED_SPECS = {}
        with open(r"{}\\test_data.json".format(OUTPUT_DIR), "r", encoding="utf-8") as f:
            test_data = json.load(f)
        for item in selected_led_cols:
            for data in test_data:
                if item.lower() == data["reference"].lower():
                    pnum = data["pnum"]
                    self.pnum_item_map.update({"PNUM-" +pnum:  item})
                    self.LED_SPECS.update(
                        {
                            item: {
                                "lsl": data["min"],
                                "usl": data["max"],
                                "pnum":pnum,
                                "unit": data["unit"],
                                "nominal" : (float(data["min"]) + float(data["max"])) / 2,
                                "tol" : round((float(data["max"]) - float(data["min"])), 3),
                            }
                        }
                    )

        self.pnum_list= sorted(self.pnum_item_map)
        for pnum in self.pnum_list:
            for data in test_data:
                if pnum == data["pnum"]:
                    item = data["reference"]
                    self.LED_SPECS.update(
                        {
                            item: {
                                "lsl": data["min"],
                                "usl": data["max"],
                                "pnum":data["pnum"],
                                "unit": data["unit"],
                                "nominal" : (float(data["min"]) + float(data["max"])) / 2,
                                "tol" : round((float(data["max"]) - float(data["min"])), 3),
                            }
                        }
                    )
        logger.info("simon select pnum_item_map \n {}".format(self.pnum_item_map))
        logger.info("simon select pnum_list \n {}".format(self.pnum_list))
        logger.info("simon select self LED_SPECS \n {}".format(self.LED_SPECS))
        # Data rows (rows 11–100): 3 inspectors × 30 trials
        # Layout matches Minitab SET C1/C2/C3 — no aggregation, raw reps:
        #   SET C2 (inspector) cycles 30 times per inspector (in inspector_numbers order)
        #   SET C1 (sample)      cycles 1–30 within each inspector block
        #   SET C3 (value)       is the raw LED reading for that (part, trial)
        #
        # Resulting row layout:
        #   rows 11–40   → inspector[0], sample 1–30 (parts 1–10, trials 1–3 each)
        #   rows 41–70   → inspector[1], sample 1–30
        #   rows 71–100  → inspector[2], sample 1–30
        #
        # PNUM columns (C–K) are filled only for items actually selected.
        # Unselected PNUM columns are left empty (template was cleared).
        #
        # Note: we do NOT filter df by inspector — work-order IDs and
        # appraiser IDs may not match. The Minitab Set C2 block order
        # comes from the template (C5/C6/C7) or config; the data we have
        # is the union of all measurements, used as the SET C3 pool.
        try:
            sn_col = next((c for c in df.columns if c.lower() in ("part_num", "sn", "sample")), None)
            if sn_col is None:
                logger.warning("GRR Form: sn column not detected in df")
                return

            df2 = df.copy()
            df2["_sn"] = df2[sn_col].astype(str)
            trial_col = next((c for c in df.columns if c.lower() == "trial"), None)
            if trial_col:
                df2["_trial"] = df2[trial_col].astype(int)
            else:
                sort_keys = ["_sn"]
                if "starttime" in df2.columns:
                    sort_keys.append("starttime")
                df2 = df2.sort_values(sort_keys).reset_index(drop=True)
                df2["_trial"] = df2.groupby("_sn").cumcount() + 1

            # Sort by (part_num, trial) — this becomes the SET C3 ordering
            # for every inspector block (Set C2 doesn't reorder the data).
            df2 = df2.sort_values(["_sn", "_trial"]).reset_index(drop=True)

            # Map each (part, trial) → {pnum: led_value}
            pnum_order = [f"PNUM-{p}" for p in list(self.pnum_item_map.keys())]
            # pnum_order = [f"PNUM-{p}" for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]

            # Order inspectors by inspector_numbers (Set C2 logic)
            insp_order = [str(x) for x in inspector_numbers[:3]]
            if not insp_order:
                logger.warning("GRR Form: no inspector_numbers — cannot fill Inspector column")
                return

            # Clear any stale values in rows 11-100 first
            for r in range(11, 101):
                for c in range(1, 12):
                    ws.cell(r, c).value = None

            # Available PNUMs (in user's selection order) — only those
            # whose LED column is in df AND the LED is in the user's
            # selected items. The columns are then laid out CONTIGUOUSLY
            # in C, D, E, … (no empty slots between selected PNUMs).
            available_pnums = list(self.pnum_item_map.keys())
            # if selected_led_cols:
            #     # Preserve user selection order: walk selected_led_cols and
            #     # map back to their PNUM string. This keeps the report
            #     # ordered exactly the way the user picked the items.
            #     for led in selected_led_cols:
            #         for pnum, pn_led in PNUM_MAPPING.items():
            #             if pn_led == led and pn_led in df.columns:
            #                 if pnum not in available_pnums:
            #                     available_pnums.append(pnum)
            #                 break
            # else:
            #     for pnum in pnum_order:
            #         led = PNUM_MAPPING.get(pnum)
            #         if led and led in df.columns:
            #             available_pnums.append(pnum)

            # Row 10 (cols C–K): write selected PNUMs contiguously starting
            # at C. Unused trailing columns are cleared.
            for j in range(9):
                col = 3 + j
                if j < len(available_pnums):
                    ws.cell(10, col).value = available_pnums[j]
                else:
                    ws.cell(10, col).value = None

            # Build the 30-row sequence for one inspector block.
            #   sample pattern: 1, 2, 3, …, 10, 1, 2, 3, …, 10, 1, 2, 3, …, 10
            #   (i.e. the 1..n_parts range repeats 3 times per block, 9 times
            #   overall across the 3 inspector blocks)
            # The value at row s_idx inside a block is (s_idx % n_parts) + 1.
            n_trials_cfg = int(GRR_FORM_DEFAULTS.get("n_trials", 3) or 3)
            n_parts_cfg = int(GRR_FORM_DEFAULTS.get("n_parts", 10) or 10)
            rows_per_block = n_parts_cfg * n_trials_cfg
            block_seq = []
            for _, drow in df2.iterrows():
                block_seq.append(drow)
            # If df has fewer than rows_per_block reps, pad with empty
            while len(block_seq) < rows_per_block:
                block_seq.append(None)

            for op_idx in range(min(3, len(insp_order))):
                op_id = insp_order[op_idx]
                for s_idx in range(rows_per_block):
                    r = 11 + op_idx * rows_per_block + s_idx
                    if r > 100:
                        break
                    drow = block_seq[s_idx]
                    # Sample number cycles 1..n_parts across the block
                    ws.cell(r, 1).value = (s_idx % n_parts_cfg) + 1
                    ws.cell(r, 2).value = str(op_id)
                    if drow is None:
                        continue
                    # PNUM columns laid out contiguously: j-th selected
                    # PNUM goes to column (3 + j). Values pulled directly
                    # from the (part, trial) row in df2.
                    logger.info("simon available_pnums is \n {} \n self.pnum_item_map is {}".format(available_pnums, self.pnum_item_map))
                    for j, pnum in enumerate(available_pnums):
                        # led = PNUM_MAPPING.get(pnum)
                        led = self.pnum_item_map.get(pnum)
                        if led and led in df.columns:
                            val = drow.get(led, None)
                            logger.info("simon value is \n {}".format(val))
                            if pd.notna(val):
                                try:
                                    # ws.cell(r, 3 + j).value = int(round(float(val)))
                                    ws.cell(r, 3 + j).value = float(val)
                                except (ValueError, TypeError):
                                    pass
        except Exception as e:
            logger.warning("GRR Form data update failed: %s", e)

    def _update_summary_page(self, ws, results, grr_items, product_name, station, project_name):
        """Update the EXISTING template Summary sheet in place.
        Template layout (rows 1–10):
          Row 1: headers (Product | Station | Pnum | Description | LowLimit | UpLimit | Unit | Tolerance | %GR&| %PT | NDC)
          Rows 2–10: per-item values
        """
        logger.info("simon grr_items is \n {}".format(grr_items))
        for i, item in enumerate(grr_items, start=2):
            if i > 10:  # template cap
                break
            grr_res = results[item].get("grr") if item in results else None
            spec = self.LED_SPECS.get(item, {})
            pnum_raw = spec.get("pnum", item[-4:])
            pnum_str = str(pnum_raw) if not isinstance(pnum_raw, str) else pnum_raw
            led = item

            # Column A: Product
            ws.cell(i, 1).value = product_name
            # Column B: Station
            ws.cell(i, 2).value = " " + station
            # Column C: Pnum (preserve original type — str for "Summary", int for "Summary ")
            existing_c = ws.cell(i, 3).value
            if isinstance(existing_c, int):
                try:
                    ws.cell(i, 3).value = int(pnum_str)
                except (ValueError, TypeError):
                    ws.cell(i, 3).value = pnum_str
            else:
                ws.cell(i, 3).value = pnum_str
            # Column D: Description
            ws.cell(i, 4).value = led
            # Column E: LowLimit
            lsl = spec.get("lsl", "")
            ws.cell(i, 5).value = lsl
            # Column F: UpLimit
            usl = spec.get("usl", "")
            ws.cell(i, 6).value = usl
            # Column G: Unit
            if len(GRR_FORM_DEFAULTS["measurement_unit"]) < 2:
                unit = spec.get("unit", "")
            else:
                unit = GRR_FORM_DEFAULTS["measurement_unit"]
            ws.cell(i, 7).value = unit
            # ws.cell(i, 7).value = GRR_FORM_DEFAULTS["measurement_unit"]
            # Column H: Tolerance (=F-E)
            if isinstance(lsl, (int, float)) and isinstance(usl, (int, float)):
                ws.cell(i, 8).value = usl - lsl
            # Column I: %GRR
            if grr_res:
                ws.cell(i, 9).value = round(float(grr_res.grr_pct), 2) if grr_res.grr_pct else 0
                ws.cell(i, 10).value = round(float(grr_res.pt_ratio), 2) if grr_res.pt_ratio else 0
                ws.cell(i, 11).value = round(float(grr_res.ndc), 2) if grr_res.ndc else 0
                # V257: Result column (PASS/FAIL) derived from
                # the Minitab GRR grade. AIAG MSA threshold: a
                # %GRR < 30% is considered acceptable for
                # production use; >= 30% is not acceptable.
                _pct_v = float(grr_res.grr_pct) if grr_res.grr_pct else 0
                ws.cell(i, 12).value = "PASS" if _pct_v < 30 else "FAIL"
            else:
                ws.cell(i, 9).value = 0
                ws.cell(i, 10).value = 0
                ws.cell(i, 11).value = 0
                ws.cell(i, 12).value = "FAIL"

    # ── Summary Page ──────────────────────────────

    def _make_summary_page(self, wb, results: Dict, grr_items: List[str]):
        ws = wb.create_sheet("Summary")

        # Column widths
        widths = [18, 10, 8, 30, 12, 12, 10, 14, 10, 10, 10, 16, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Title
        ws.merge_cells("A1:M1")
        _apply_header(ws.cell(1, 1), "Gage R&R Summary Report", bg=CLR_HEADER_BG, size=12, bold=True)
        ws.row_dimensions[1].height = 24

        # Column headers
        headers = [
            "Product",
            "Station",
            "pnum",
            "Description",
            "LowLimit",
            "UpLimit",
            "Unit",
            "Tolerance",
            "%GRR",
            "%PT",
            "NDC",
            "GR&R Result",
            "Remark",
        ]
        for j, h in enumerate(headers, start=1):
            _apply_header(ws.cell(2, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, wrap=True)
        ws.row_dimensions[2].height = 30

        # Data rows
        for i, item in enumerate(grr_items):
            r = i + 3
            grr = results[item].get("grr")
            spec = LED_SPECS.get(item, {})
            lsl = spec.get("lsl", 0)
            usl = spec.get("usl", 65555)
            tol = usl - lsl

            pct_grr = grr.grr_pct if grr else 0.0
            pct_pt = grr.pt_ratio if grr else 0.0
            ndc = grr.ndc if grr else 0

            res_bg, res_fg, res_text = _grr_result_color(pct_grr)
            pass_fail = _result_pass_fail(pct_grr)

            bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"

            _apply_data(ws.cell(r, 1), "Beta Charger", bg=bg)
            _apply_data(ws.cell(r, 2), "MT7", bg=bg)
            pnum_str = LED_SPECS.get(item, {}).get("pnum", item[-4:])
            _apply_data(ws.cell(r, 3), pnum_str, bg=bg, fmt="0")
            _apply_data(ws.cell(r, 4), item, bg=bg, align="left")
            _apply_data(ws.cell(r, 5), lsl, bg=bg, fmt="#,##0")
            _apply_data(ws.cell(r, 6), usl, bg=bg, fmt="#,##0")
            _apply_data(ws.cell(r, 7), "Lumen", bg=bg)
            _apply_data(ws.cell(r, 8), tol, bg=bg, fmt="#,##0")
            _apply_data(ws.cell(r, 9), round(pct_grr, 2), bg=bg, fmt="0.00")
            _apply_data(ws.cell(r, 10), round(pct_pt, 2), bg=bg, fmt="0.00")
            _apply_data(ws.cell(r, 11), ndc, bg=bg, fmt="0")

            # GR&R Result with colour
            cell_res = ws.cell(r, 12)
            cell_res.value = pass_fail
            cell_res.font = Font(name="Calibri", bold=True, color=res_fg, size=10)
            cell_res.fill = _hdr_fill(res_bg)
            cell_res.border = _thin_border()
            cell_res.alignment = Alignment(horizontal="center", vertical="center")

            _apply_data(ws.cell(r, 13), "", bg=bg)

        # Notes row
        note_row = len(grr_items) + 4
        ws.merge_cells(f"A{note_row}:M{note_row}")
        note = ws.cell(
            note_row,
            1,
            value="a.%GR&R <=10%, Good;  10%<%GR&R<=30% Acceptable; %GR&R>30% Bad. "
            "b.%P/T<=10% Good;  10%<%P/T<=30% Acceptable; %P/T>30% Bad. "
            "c.NDC>10 Good; 5<=NDC<=10 Acceptable, NDC<5 Bad",
        )
        note.font = Font(name="Calibri", italic=True, size=8, color="595959")
        note.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[note_row].height = 40

        ws.freeze_panes = "A3"

    # ── Per-PNUM Sheet ────────────────────────────

    def _make_pnum_page(self, wb, item: str, grr_res, spec: dict, chart_paths: Optional[dict] = None):
        """Create a sheet for one PNUM item with GRR metrics and embedded chart(s).
        Sheet name follows GRR_SHEET_NAMING (default: just the 4-digit Pnum, e.g. "4024").

        chart_paths: dict mapping chart key ('grr_<safe>', 'grr_<safe>_xbar', 'grr_<safe>_r')
                     to PNG file path. The function picks the first two available
                     charts and embeds them at the positions in GRR_CHART_LAYOUT.
        """
        chart_paths = chart_paths or {}
        pnum_raw = LED_SPECS.get(item, {}).get("pnum", item[-4:])
        pnum_str = str(pnum_raw) if not isinstance(pnum_raw, str) else pnum_raw
        if GRR_SHEET_NAMING == "pnum":
            sheet_name = pnum_str
        else:
            sheet_name = f"PNUM-{pnum_str}"
        # Excel limit: 31 chars max
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 20

        # ── Row 1: section headers ──
        ws.merge_cells("A1:E1")
        _apply_header(ws.cell(1, 1), "1.Gage R&R Data", bg=CLR_HEADER_BG)
        ws.merge_cells("F1:G1")
        _apply_header(ws.cell(1, 6), "2.Graph Report", bg=CLR_HEADER_BG)

        # ── Row 2: sub-headers ──
        ws.merge_cells("A2:E2")
        _apply_header(ws.cell(2, 1), "3.Conclusion:", bg=CLR_SECTION_BG, fg="1F4E79", align="left")

        # ── Metrics rows ──
        pct_grr = grr_res.grr_pct if grr_res else 0.0
        pct_pt = grr_res.pt_ratio if grr_res else 0.0
        ndc = grr_res.ndc if grr_res else 0

        res_bg, res_fg, res_text = _grr_result_color(pct_grr)

        lsl = spec.get("lsl", 0)
        usl = spec.get("usl", 65555)

        metrics = [
            (None, None, None, None, None),
            (
                "%GR&R=",
                round(pct_grr, 2),
                None,
                None,
                "a.%GR&R <=10%, Good;  10%<%GR&R<=30% Acceptable; %GR&R>30% Bad.",
            ),
            ("%P/T=", round(pct_pt, 2), None, None, "b.%P/T<=10% Good;  10%<%P/T<=30% Acceptable; %P/T>30% Bad."),
            ("NDC=", ndc, None, None, "c.NDC>10 Good; 5<=NDC<=10 Acceptable, NDC<5 Bad"),
            (None, None, None, None, None),
            ("LSL=", lsl, None, None, None),
            ("USL=", usl, None, None, None),
            ("Tolerance=", usl - lsl, None, None, None),
        ]

        for i, row_data in enumerate(metrics, start=3):
            for j, val in enumerate(row_data, start=1):
                if val is None:
                    ws.cell(i, j).border = _thin_border()
                    continue
                is_metric_label = j == 1 and val and "=" in str(val)
                is_metric_val = j == 2 and i in (4, 5, 6)
                is_note = j == 5

                if is_note:
                    ws.merge_cells(f"E{i}:G{i}")
                    cell = ws.cell(i, j)
                    cell.value = str(val)
                    cell.font = Font(name="Calibri", size=8, color="595959")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = _thin_border()
                    ws.row_dimensions[i].height = 20
                elif is_metric_label:
                    cell = ws.cell(i, j)
                    cell.value = str(val)
                    cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = _thin_border()
                elif is_metric_val and i == 4:  # %GR&R value
                    cell = ws.cell(i, j)
                    cell.value = float(val) / 100 if isinstance(val, (int, float)) and val > 1 else val
                    cell.number_format = "0.00%"
                    cell.font = Font(name="Calibri", bold=True, size=11, color=res_fg)
                    cell.fill = _hdr_fill(res_bg)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif is_metric_val and i == 5:  # %P/T value
                    cell = ws.cell(i, j)
                    cell.value = float(val) / 100 if isinstance(val, (int, float)) and val > 1 else val
                    cell.number_format = "0.00%"
                    cell.font = Font(name="Calibri", bold=True, size=11)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif is_metric_val and i == 6:  # NDC value
                    cell = ws.cell(i, j)
                    cell.value = val
                    cell.font = Font(name="Calibri", bold=True, size=11, color="1F4E79" if ndc >= 10 else "9C6500")
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell = ws.cell(i, j)
                    cell.value = val
                    cell.font = Font(name="Calibri", size=10)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        cell.number_format = "#,##0"

        # ── Result row ──
        res_row = 11
        ws.merge_cells(f"A{res_row}:D{res_row}")
        _apply_header(ws.cell(res_row, 1), "GR&R Result:", bg=res_bg, fg=res_fg, align="left")
        ws.merge_cells(f"E{res_row}:G{res_row}")
        cell_result = ws.cell(res_row, 5)
        cell_result.value = res_text
        cell_result.font = Font(name="Calibri", bold=True, size=11, color=res_fg)
        cell_result.fill = _hdr_fill(res_bg)
        cell_result.border = _thin_border()
        cell_result.alignment = Alignment(horizontal="center", vertical="center")

        # ── Graph Report: embed images via OneCellAnchor (same method as ARR) ──
        # chart_paths keys are now prefixed with the item's safe name to avoid
        # cross-item collisions: "<safe>__slideNN_shapeNN". Pick the first two
        # available charts (sorted by slide/shape order) for left/right slots.
        safe = item.replace("/", "_")
        item_chart_keys = [k for k in (chart_paths or {}).keys() if k.startswith(safe + "__")]
        item_chart_keys.sort()  # slide01_shape01 < slide01_shape02 < slide02_shape01
        chosen = []
        for k in item_chart_keys[:2]:
            v = chart_paths[k]
            if v and os.path.isfile(v):
                chosen.append((k.split("__")[-1], v))

        if chosen:
            try:
                # Clear any template images first (defensive — keeps behaviour
                # consistent with ARR's AR&R Report handling).
                if hasattr(ws, "_images") and ws._images:
                    logger.info("Clearing %d pre-existing images on sheet '%s'", len(ws._images), sheet_name)
                    ws._images.clear()

                from openpyxl.drawing.image import Image as _Img
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPositiveSize2D

                # If only one chart available, use the 'main' slot
                layout_keys = list(GRR_CHART_LAYOUT.keys())
                for idx, (chart_key, chart_src) in enumerate(chosen):
                    layout_key = layout_keys[idx] if idx < len(layout_keys) else layout_keys[-1]
                    col, row, w, h, colOff_px, rowOff_px = GRR_CHART_LAYOUT[layout_key]

                    # chart_src may be a disk path (legacy) or a BytesIO
                    # buffer (in-memory per-item cleanup in
                    # GRRAnalyzer.run_all_minitab). Both work with openpyxl.
                    if hasattr(chart_src, "read"):
                        # BytesIO: rewind to start, Image() will read from
                        # current position
                        try:
                            chart_src.seek(0)
                        except Exception:
                            pass
                        img = _Img(chart_src)
                    else:
                        img = _Img(chart_src)
                    img.width = w
                    img.height = h
                    # 1 px = 9525 EMU at 96 DPI; ext is required or openpyxl
                    # silently drops the image.
                    anchor = OneCellAnchor(
                        _from=AnchorMarker(
                            col=col - 1,
                            colOff=int(colOff_px * 9525),
                            row=row - 1,
                            rowOff=int(rowOff_px * 9525),
                        ),
                        ext=XDRPositiveSize2D(cx=int(w * 9525), cy=int(h * 9525)),
                    )
                    img.anchor = anchor
                    ws.add_image(img)
                    logger.info(
                        "GRR chart #%d ('%s'/%s) embedded on sheet '%s' at R%dC%d (%dx%d)",
                        idx + 1,
                        getattr(chart_src, "name", os.path.basename(str(chart_src))),
                        layout_key,
                        sheet_name,
                        row,
                        col,
                        w,
                        h,
                    )
            except Exception as eg:
                logger.warning("Could not embed GRR charts on '%s': %s", sheet_name, eg)
        else:
            # Placeholder text in graph area
            ws.merge_cells("F2:G10")
            cell_g = ws.cell(2, 6)
            cell_g.value = "(Chart: GageRR graph)"
            cell_g.font = Font(name="Calibri", size=9, color="AAAAAA", italic=True)
            cell_g.alignment = Alignment(horizontal="center", vertical="center")

        # ── EV / PV / TV detail table ──
        detail_row = 13
        ws.merge_cells(f"A{detail_row}:G{detail_row}")
        _apply_header(
            ws.cell(detail_row, 1), "GRR Component Analysis", bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, align="left"
        )

        comp_headers = ["Component", "Value", "Unit", "", "", "", ""]
        for j, h in enumerate(comp_headers, start=1):
            _apply_header(ws.cell(detail_row + 1, j), h, bg=CLR_SECTION_BG, fg="1F4E79", bold=True)

        if grr_res:
            comp_data = [
                ("EV (Repeatability)", round(grr_res.ev, 2) if grr_res.ev else "-", "Lumen"),
                ("PV (Reproducibility)", round(grr_res.pv, 2) if grr_res.pv else "-", "Lumen"),
                ("TV (Total Variation)", round(grr_res.tv, 2) if grr_res.tv else "-", "Lumen"),
                ("GR&R", round(grr_res.tv, 2) if grr_res.tv else "-", "Lumen"),
                ("%GR&R", round(pct_grr, 2), "%"),
                ("%P/T", round(pct_pt, 2), "%"),
                ("NDC", ndc, "-"),
            ]
        else:
            comp_data = []

        for i2, (comp, val, unit) in enumerate(comp_data, start=detail_row + 2):
            bg = CLR_ALT_ROW if i2 % 2 == 0 else "FFFFFF"
            _apply_data(ws.cell(i2, 1), comp, bg=bg, align="left")
            cell_v = ws.cell(i2, 2)
            cell_v.value = val
            cell_v.font = Font(name="Calibri", bold=True, size=10)
            cell_v.fill = _hdr_fill(bg)
            cell_v.border = _thin_border()
            cell_v.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(val, float):
                cell_v.number_format = "0.00"
            _apply_data(ws.cell(i2, 3), unit, bg=bg)

        # ── Footer ──
        footer_row = detail_row + 2 + len(comp_data) + 1
        ws.merge_cells(f"A{footer_row}:G{footer_row}")
        _apply_data(ws.cell(footer_row, 1), "This Template Link to #WI-002388", align="left", bold=True, fg="595959")

        ws.freeze_panes = "A2"

    def _update_pnum_page(self, ws, item: str, grr_res, spec: dict, chart_paths: Optional[dict] = None):
        """Update an EXISTING per-item template sheet (PNUM-XXXX or 4-digit).
        - Replaces F24/F25/F26 with the new %GR&R / %P/T / NDC values.
        - Clears any old images, then embeds new Minitab charts at the
          positions defined by GRR_LAYOUT (mirrors the AR&R pattern).
        """
        pct_grr = grr_res.grr_pct if grr_res else 0.0
        pct_pt = grr_res.pt_ratio if grr_res else 0.0
        ndc = grr_res.ndc if grr_res else 0

        # Update the three GRR metric cells in the template layout:
        #   D24 = '%GR&R=' label, F24 = value
        #   D25 = '%P/T=' label,  F25 = value
        #   D26 = 'NDC=' label,   F26 = value
        ws.cell(24, 6).value = round(float(pct_grr), 2) if pct_grr else 0.0
        ws.cell(25, 6).value = round(float(pct_pt), 2) if pct_pt else 0.0
        ws.cell(26, 6).value = round(float(ndc), 2) if ndc else 0

        # Clear old images (template may have its own placeholders)
        if hasattr(ws, "_images") and ws._images:
            logger.info("Clearing %d old images on template sheet '%s'", len(ws._images), ws.title)
            ws._images.clear()

        # ── Embed Minitab GRR charts (full PPT-derived image set) ──
        # chart_paths comes from GRRAnalyzer._run_minitab: keys are
        # image stems (e.g. 'slide01_shape02') and values are PNG paths.
        # _img_dir / _extracted are internal metadata; skip them.
        chart_paths = chart_paths or {}
        if not chart_paths:
            return

        try:
            from openpyxl.drawing.image import Image as _Img
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D

            img_dir = chart_paths.get("_img_dir", "")
            placed = 0
            for img_name, col, row, w, h, colOff_px, rowOff_px in GRR_LAYOUT:
                # Look up the chart by stem (filename minus extension)
                stem = Path(img_name).stem
                chart_src = chart_paths.get(stem)
                if chart_src is None:
                    # Also try the full filename (legacy disk path fallback)
                    chart_src = str(Path(img_dir) / img_name) if img_dir else img_name
                # chart_src may be a BytesIO buffer (in-memory cleanup) or
                # a disk path. Both are accepted.
                if hasattr(chart_src, "read"):
                    try:
                        chart_src.seek(0)
                    except Exception:
                        pass
                elif not Path(str(chart_src)).exists():
                    logger.warning("GRR chart '%s' not found, skipping", img_name)
                    continue

                img = _Img(chart_src)
                img.width = w
                img.height = h
                anchor = OneCellAnchor(
                    _from=AnchorMarker(
                        col=col - 1,
                        colOff=int(colOff_px * 9525),
                        row=row - 1,
                        rowOff=int(rowOff_px * 9525),
                    ),
                    ext=XDRPositiveSize2D(cx=int(w * 9525), cy=int(h * 9525)),
                )
                img.anchor = anchor
                ws.add_image(img)
                placed += 1
            if placed:
                logger.info("GRR sheet '%s': embedded %d chart(s) via GRR_LAYOUT", ws.title, placed)
        except Exception as e:
            logger.warning("Failed to embed GRR charts on '%s': %s", ws.title, e)


# Reverse mapping for LED → PNUM
PNUM_MAPPING_REV = {v: k for k, v in PNUM_MAPPING.items()}
