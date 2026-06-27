# ──────────────────────────────────────────────
#  Excel Report Generator
#  Produces a FORM-004090 GRR Excel report
# ──────────────────────────────────────────────
import logging
from datetime import datetime
import os
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from utils.config import LED_SPECS, PNUM_MAPPING, _norm_led

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────
CLR_HEADER_BG   = "1F4E79"   # dark blue
CLR_HEADER_FG   = "FFFFFF"
CLR_SUBHDR_BG   = "2E75B6"
CLR_SUBHDR_FG   = "FFFFFF"
CLR_SECTION_BG  = "D6E4F0"
CLR_PASS_BG     = "E2EFDA"   # light green
CLR_FAIL_BG     = "FCE4D6"   # light red
CLR_WARN_BG     = "FFF2CC"   # light yellow
CLR_ALT_ROW     = "EBF3FB"   # alternating row
CLR_NOTE_BG     = "FFFACD"
CLR_GRR_GOOD    = "C6EFCE"
CLR_GRR_ACC     = "FFEB9C"
CLR_GRR_BAD     = "FFC7CE"

# ── Helper styles ───────────────────────────────
def _hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _bold_border() -> Border:
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header(cell, text: str, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG, size=10, bold=True, wrap=False, align="center"):
    cell.value = text
    cell.font  = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill  = _hdr_fill(bg)
    cell.border= _thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def _apply_data(cell, value, fmt=None, bold=False, align="center", bg=None, fg="000000"):
    cell.value = value
    cell.font  = Font(name="Calibri", bold=bold, color=fg, size=10)
    if bg:
        cell.fill = _hdr_fill(bg)
    cell.border    = _thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt

def _grr_result_color(pct_grr: float) -> tuple:
    if pct_grr <= 10:
        return CLR_GRR_GOOD, "000000", "Good"
    elif pct_grr <= 30:
        return CLR_GRR_ACC, "9C6500", "Acceptable"
    else:
        return CLR_GRR_BAD, "9C0006", "Bad"

def _result_pass_fail(pct_grr: float) -> str:
    if pct_grr <= 30:
        return "PASS"
    return "FAIL"

# ──────────────────────────────────────────────
#  Main Report Generator
# ──────────────────────────────────────────────
class ExcelReportGenerator:
    """
    Generates a FORM-004090 GRR Excel report matching the company template.

    Sheet structure:
      1. Cover Page   – approvals + revision history
      2. GRR From     – raw measurement data (Sample, Inspector, PNUM-4024~PNUM-4056)
      3. Summary       – %GRR / %PT / NDC results for all parameters
      4. PNUM-4024    – per-parameter GRR detail + graph placeholder
      ...
      N. PNUM-4056
    """

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, results: Dict, df: pd.DataFrame,
                 inline_charts: bool = False,
                 minitab_path: Optional[str] = None,
                 product_name: str = "Hermes Min Charger",
                 station: str = "MT7",
                 project_name: str = "Beta",
                 reported_by: str = "Min Duan",
                 inspectors: Optional[List[str]] = None) -> str:
        """
        Generate the full Excel GRR report.

        Args:
            results:       {item_name: {"grr": GRRResult, "cpk": CPKResult}}
            df:            Full GRR DataFrame (Sample, Inspector, PNUM cols)
            inline_charts: Whether to embed Minitab charts
            minitab_path:  Path to Minitab Mtb.exe
            product_name:  Product name for report header
            station:       Station ID (e.g. MT7)
            project_name:  Project name
            reported_by:   Who generated the report
            inspectors:    List of inspector numbers (auto-detected if None)

        Returns:
            Path to the saved .xlsx file
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        # ── Gather metadata ──
        grr_items = [k for k, v in results.items() if "grr" in v and v["grr"]]
        if not inspectors:
            # Auto-detect inspectors from df
            insp_col = [c for c in df.columns if c.lower() in ("inspector", "appraiser")][0]
            inspectors = sorted(df[insp_col].dropna().unique().tolist(), key=str)
        inspector_numbers = [str(i) for i in inspectors]

        date_str = datetime.now().strftime("%B %d, %Y")
        ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")

        # ── 1. Cover Page ──
        self._make_cover_page(wb, product_name, station, project_name,
                              reported_by, inspector_numbers, date_str)

        # ── 2. GRR From (raw data) ──
        self._make_grr_data_page(wb, df, inspector_numbers)

        # ── 3. Summary ──
        self._make_summary_page(wb, results, grr_items)

        # ── 4. Per-parameter sheets ──
        for item in grr_items:
            grr_res = results[item]["grr"]
            spec = LED_SPECS.get(item, {})
            chart_path = grr_res.chart_paths.get('grr_{}'.format(item.replace('/', '_'))
                             ) if hasattr(grr_res, 'chart_paths') else None
            self._make_pnum_page(wb, item, grr_res, spec, chart_path)

        # ── Save ──
        filename = f"Beta-GRR-Charger_{station}-HVTE-M600099-{datetime.now():%Y%m%d}.xlsx"
        out_path = self.output_dir / filename
        wb.save(out_path)
        logger.info(f"GRR Excel report saved: {out_path}")
        return str(out_path)

    # ── Cover Page ────────────────────────────────

    def _make_cover_page(self, wb, product, station, project,
                          reported_by, inspectors, date_str):
        ws = wb.create_sheet("Cover Page")
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20

        # Row 1-2: blank (spacer)
        # Row 3-5: Approvals header
        for r in range(3, 11):
            ws.row_dimensions[r].height = 18

        _apply_header(ws.cell(3, 1), "APPROVALS", bg=CLR_HEADER_BG, size=11, wrap=True)
        _apply_header(ws.cell(3, 2), "DEPT",      bg=CLR_HEADER_BG)
        _apply_header(ws.cell(3, 3), "NAME",      bg=CLR_HEADER_BG)
        _apply_header(ws.cell(3, 4), "TITLE",     bg=CLR_HEADER_BG)
        _apply_header(ws.cell(3, 5), "DATE",      bg=CLR_HEADER_BG)

        approval_data = [
            ("QA", "Kerwin Wang", "Asst. Manager", "In EQMS"),
        ]
        for i, row in enumerate(approval_data, start=4):
            for j, val in enumerate(row, start=2):
                _apply_data(ws.cell(i, j), val, align="center")

        # Approval rows (blank)
        for r in range(5, 11):
            for c in range(1, 6):
                ws.cell(r, c).border = _thin_border()

        # ── Title block (columns F-G merged) ──
        ws.merge_cells("F3:G3")
        _apply_header(ws.cell(3, 6), "TEMPLATE TITLE", bg=CLR_HEADER_BG, size=11, wrap=True)

        ws.merge_cells("F4:G4")
        _apply_header(ws.cell(4, 6), "Gage R and R Form", bg=CLR_SUBHDR_BG, size=12, bold=True)

        ws.merge_cells("F5:G5")
        _apply_header(ws.cell(5, 6), "Gage R and R Form", bg=CLR_SUBHDR_BG, size=12, bold=True)

        ws.merge_cells("F6:G6")
        _apply_header(ws.cell(6, 6), "TEMPLATE NUMBER FORM-004090", bg=CLR_SUBHDR_BG, size=10)

        ws.merge_cells("F7:G7")
        _apply_header(ws.cell(7, 6), "REVISION B", bg=CLR_SUBHDR_BG, size=10)

        # ── Revision history ──
        ws.merge_cells("A10:G10")
        _apply_header(ws.cell(10, 1), "REVISION HISTORY", bg=CLR_HEADER_BG, size=10, wrap=True)

        rev_headers = ["REVISION", "DESCRIPTION OF CHANGE", "ORIGINATOR", "RELEASE DATE"]
        for j, h in enumerate(rev_headers, start=1):
            _apply_header(ws.cell(11, j), h, bg=CLR_SUBHDR_BG)
        for j in range(5, 8):
            ws.cell(11, j).border = _thin_border()

        rev_data = [
            ("A", "Transfer from QDMS to EQMS", "Li Rao", "In EQMS"),
            ("B", "Update header and footer for the template", "Li Rao", "In EQMS"),
        ]
        for i, row in enumerate(rev_data, start=12):
            for j, val in enumerate(row, start=1):
                _apply_data(ws.cell(i, j), val, align="left" if j == 2 else "center")
            for j in range(5, 8):
                ws.cell(i, j).border = _thin_border()

        # Note row
        ws.merge_cells("A14:G14")
        note_cell = ws.cell(14, 1,
            value="Note: The cover page is only used for electronic file traceability of change history, "
                  "no need print out the cover page when do hardcopy record"
        )
        note_cell.font = Font(name="Calibri", italic=True, size=8, color="595959")
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[14].height = 30

        # ── Freeze pane ──
        ws.freeze_panes = "A3"

    # ── GRR Raw Data Page ────────────────────────

    def _make_grr_data_page(self, wb, df: pd.DataFrame, inspector_numbers: List[str]):
        ws = wb.create_sheet("GRR From")

        # Column widths
        ws.column_dimensions["A"].width = 8   # Sample
        ws.column_dimensions["B"].width = 12  # Inspector
        for col_idx in range(3, 12):          # PNUM columns
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # ── Header block (rows 1-11) ──
        ws.merge_cells("A1:K1")
        ws.merge_cells("A2:K2")

        ws.merge_cells("A3:C3")
        _apply_header(ws.cell(3, 1), "Part Number :", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        ws.cell(3, 4).value = "PMMGH-DAC1471"
        _apply_data(ws.cell(3, 4), "PMMGH-DAC1471", align="center")
        ws.merge_cells("D3:K3")

        ws.merge_cells("A4:C4")
        _apply_header(ws.cell(4, 1), "Instrument name:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        ws.merge_cells("D4:G4")
        _apply_data(ws.cell(4, 4), "Charger MT7 Tester", align="center")

        ws.merge_cells("H4:I4")
        _apply_header(ws.cell(4, 8), "Instrument No:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(4, 10), "HVTE-M600099", align="center")

        ws.merge_cells("A5:C5")
        _apply_header(ws.cell(5, 1), "Department Name:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(5, 4), "TE", align="center")

        ws.merge_cells("H5:I5")
        _apply_header(ws.cell(5, 8), "Reported by:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(5, 10), reported_by_val := "Min Duan", align="center")

        ws.merge_cells("A6:C6")
        _apply_header(ws.cell(6, 1), "Inspector Number:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        for i, insp in enumerate(inspector_numbers[:3]):
            _apply_data(ws.cell(6, 4 + i), insp, align="center")

        ws.merge_cells("H6:I6")
        _apply_header(ws.cell(6, 8), "Measurement Unit:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(6, 10), "Lumen", align="center")

        ws.merge_cells("A7:C7")
        _apply_header(ws.cell(7, 1), "Inspector Number:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        for i, insp in enumerate((inspector_numbers * 2)[:3]):
            _apply_data(ws.cell(7, 4 + i), insp, align="center")

        ws.merge_cells("H7:I7")
        _apply_header(ws.cell(7, 8), "Project Name:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(7, 10), "Beta", align="center")

        ws.merge_cells("A8:C8")
        _apply_header(ws.cell(8, 1), "Inspector Number:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        for i, insp in enumerate(inspector_numbers[:3]):
            _apply_data(ws.cell(8, 4 + i), insp, align="center")

        ws.merge_cells("H8:I8")
        _apply_header(ws.cell(8, 8), "Date Inspected:", bg=CLR_SECTION_BG, align="left", fg="1F4E79")
        _apply_data(ws.cell(8, 10), datetime.now().strftime("%B %d, %Y"), align="center")

        # ── Column headers (row 9) ──
        col_headers = ["Parameter"] * 2 + ["Sample", "Inspector"] + [
            f"PNUM-{p}" for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]
        ]
        for j, h in enumerate(col_headers, start=1):
            _apply_header(ws.cell(9, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, size=10)
        ws.merge_cells("A9:B9")
        _apply_header(ws.cell(9, 1), "Parameter", bg=CLR_SUBHDR_BG)
        ws.cell(9, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("C9:D9")
        _apply_header(ws.cell(9, 3), "Parameter", bg=CLR_SUBHDR_BG)

        # Row 10: column sub-headers
        _apply_header(ws.cell(10, 1), "Sample",  bg=CLR_SUBHDR_BG)
        _apply_header(ws.cell(10, 2), "Inspector", bg=CLR_SUBHDR_BG)
        for j in range(3, 12):
            ws.cell(10, j).value = None
            ws.cell(10, j).border = _thin_border()

        # ── Data rows ──
        # Determine which column holds inspector numbers
        insp_col = None
        sample_col = None
        pnum_col_map = {}
        for col in df.columns:
            cl = col.strip().lower()
            if cl in ("inspector", "appraiser"):
                insp_col = col
            elif cl == "sample":
                sample_col = col
            elif col in PNUM_MAPPING.values() or any(p in col for p in ["PNUM-", "INTENSITY"]):
                # Match PNUM column
                for pnum, led in PNUM_MAPPING.items():
                    if pnum in col or led in col:
                        pnum_col_map[pnum] = col
                        break

        # If df is already the GRR From style (PNUM-4024 style columns)
        pnum_in_df = [c for c in df.columns if c.startswith("PNUM-")]

        start_row = 11
        row_idx = start_row

        if pnum_in_df:
            # GRR template format: Sample, Inspector, PNUM-4024, ...
            pnums = ["PNUM-" + str(p) for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]
            df_out = df[[c for c in pnums if c in df.columns]]
            insp_vals = df["Inspector"].tolist() if "Inspector" in df.columns else []
            sample_vals = df["Sample"].tolist() if "Sample" in df.columns else []

            for i, (idx, row) in enumerate(df_out.iterrows()):
                bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
                sn_val = sample_vals[i] if i < len(sample_vals) else ""
                insp_val = insp_vals[i] if i < len(insp_vals) else ""

                _apply_data(ws.cell(row_idx, 1), sn_val,   bg=bg, align="center")
                _apply_data(ws.cell(row_idx, 2), insp_val, bg=bg, align="center")
                for j, pnum in enumerate(pnums):
                    val = row.get(pnum, None)
                    if pd.notna(val):
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            val = None
                    _apply_data(ws.cell(row_idx, 3 + j), val, bg=bg, align="center", fmt="0")
                row_idx += 1

        else:
            # Intermediate format: need to pivot
            # df has: sn, appraiser, LED_XXX_INTENSITY columns
            led_to_pnum = {v: k for k, v in PNUM_MAPPING.items()}

            # Determine unique parts and operators
            sn_col = [c for c in df.columns if c.lower() in ("sn", "part_num", "sample")][0] if any(
                c.lower() in ("sn", "part_num", "sample") for c in df.columns) else None
            op_col = [c for c in df.columns if c.lower() in ("appraiser", "inspector", "operator")][0] if any(
                c.lower() in ("appraiser", "inspector", "operator") for c in df.columns) else None

            if sn_col and op_col:
                df = df.copy()
                if sn_col == "part_num":
                    pass  # already numbered
                elif sn_col == "sn":
                    unique_sns = sorted(df[sn_col].dropna().unique(), key=lambda x: str(x))
                    sn_map = {s: i + 1 for i, s in enumerate(unique_sns)}
                    df["part_num"] = df[sn_col].map(sn_map)

                unique_ops = sorted(df[op_col].dropna().unique(), key=str)
                op_map = {o: str(o) for o in unique_ops}
                df["op_str"] = df[op_col].map(op_map)

                # Build GRR table rows
                led_cols = [c for c in df.columns if "INTENSITY" in c]
                for pnum in ["PNUM-" + str(p) for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]:
                    pass  # will fill below

                # Write header row for each pnum column
                pnum_headers = ["PNUM-" + str(p) for p in [4024, 4028, 4032, 4036, 4040, 4044, 4048, 4052, 4056]]
                for j, ph in enumerate(pnum_headers):
                    _apply_header(ws.cell(10, 3 + j), ph, bg=CLR_SECTION_BG, fg="1F4E79", size=9)

                row_idx = start_row
                for part in range(1, 11):
                    for op in unique_ops:
                        for trial in range(1, 4):
                            sub = df[(df["part_num"] == part) & (df["op_str"] == str(op))]
                            # Find row for this trial
                            rows_for_trial = sub.sort_values("trial").drop_duplicates("trial", keep="last")
                            for _, r in rows_for_trial.iterrows():
                                bg = CLR_ALT_ROW if (part + trial) % 2 == 0 else "FFFFFF"
                                _apply_data(ws.cell(row_idx, 1), part, bg=bg)
                                _apply_data(ws.cell(row_idx, 2), str(op), bg=bg)
                                for j, led in enumerate(led_cols):
                                    val = r.get(led, None)
                                    if pd.notna(val):
                                        try:
                                            val = float(val)
                                        except (ValueError, TypeError):
                                            val = None
                                    pnum_col = PNUM_MAPPING_REV.get(led, "")
                                    col_idx = pnum_headers.index(pnum_col) + 3 if pnum_col in pnum_headers else -1
                                    if col_idx >= 0:
                                        _apply_data(ws.cell(row_idx, col_idx), val, bg=bg, fmt="0")
                                row_idx += 1

        # ── Footer ──
        footer_row = row_idx + 1
        ws.merge_cells(f"A{footer_row}:K{footer_row}")
        _apply_data(ws.cell(footer_row, 1),
                    "This Template Link to #WI-002388",
                    align="left", bold=True, fg="595959")

        ws.freeze_panes = "A11"

    # ── Summary Page ──────────────────────────────

    def _make_summary_page(self, wb, results: Dict, grr_items: List[str]):
        ws = wb.create_sheet("Summary")

        # Column widths
        widths = [18, 10, 8, 30, 12, 12, 10, 14, 10, 10, 10, 16, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Title
        ws.merge_cells("A1:M1")
        _apply_header(ws.cell(1, 1), "Gage R&R Summary Report", bg=CLR_HEADER_BG, size=12, bold=True)
        ws.row_dimensions[1].height = 24

        # Column headers
        headers = ["Product", "Station", "pnum", "Description",
                   "LowLimit", "UpLimit", "Unit", "Tolerance",
                   "%GRR", "%PT", "NDC", "GR&R Result", "Remark"]
        for j, h in enumerate(headers, start=1):
            _apply_header(ws.cell(2, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, wrap=True)
        ws.row_dimensions[2].height = 30

        # Data rows
        for i, item in enumerate(grr_items):
            r = i + 3
            grr = results[item].get("grr")
            spec = LED_SPECS.get(item, {})
            lsl = spec.get("lsl", 0)
            usl = spec.get("usl", 65555)
            tol = usl - lsl

            pct_grr = grr.grr_pct if grr else 0.0
            pct_pt  = grr.pt_ratio  if grr else 0.0
            ndc     = grr.ndc     if grr else 0

            res_bg, res_fg, res_text = _grr_result_color(pct_grr)
            pass_fail = _result_pass_fail(pct_grr)

            bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"

            _apply_data(ws.cell(r, 1), "Beta Charger",  bg=bg)
            _apply_data(ws.cell(r, 2), "MT7",            bg=bg)
            pnum_str = LED_SPECS.get(item, {}).get("pnum", item[-4:])
            _apply_data(ws.cell(r, 3), pnum_str,         bg=bg, fmt="0")
            _apply_data(ws.cell(r, 4), item,             bg=bg, align="left")
            _apply_data(ws.cell(r, 5), lsl,              bg=bg, fmt="#,##0")
            _apply_data(ws.cell(r, 6), usl,              bg=bg, fmt="#,##0")
            _apply_data(ws.cell(r, 7), "Lumen",          bg=bg)
            _apply_data(ws.cell(r, 8), tol,               bg=bg, fmt="#,##0")
            _apply_data(ws.cell(r, 9), round(pct_grr, 2), bg=bg, fmt="0.00")
            _apply_data(ws.cell(r, 10), round(pct_pt, 2), bg=bg, fmt="0.00")
            _apply_data(ws.cell(r, 11), ndc,             bg=bg, fmt="0")

            # GR&R Result with colour
            cell_res = ws.cell(r, 12)
            cell_res.value = pass_fail
            cell_res.font  = Font(name="Calibri", bold=True, color=res_fg, size=10)
            cell_res.fill  = _hdr_fill(res_bg)
            cell_res.border= _thin_border()
            cell_res.alignment = Alignment(horizontal="center", vertical="center")

            _apply_data(ws.cell(r, 13), "", bg=bg)

        # Notes row
        note_row = len(grr_items) + 4
        ws.merge_cells(f"A{note_row}:M{note_row}")
        note = ws.cell(note_row, 1,
            value="a.%GR&R <=10%, Good;  10%<%GR&R<=30% Acceptable; %GR&R>30% Bad. "
                  "b.%P/T<=10% Good;  10%<%P/T<=30% Acceptable; %P/T>30% Bad. "
                  "c.NDC>10 Good; 5<=NDC<=10 Acceptable, NDC<5 Bad")
        note.font = Font(name="Calibri", italic=True, size=8, color="595959")
        note.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[note_row].height = 40

        ws.freeze_panes = "A3"

    # ── Per-PNUM Sheet ────────────────────────────

    def _make_pnum_page(self, wb, item: str, grr_res,
                         spec: dict, chart_img_path: Optional[str] = None):
        """Create a PNUM-XXXX sheet with GRR metrics and graph placeholder."""
        pnum_str = LED_SPECS.get(item, {}).get("pnum", item[-4:])
        sheet_name = f"PNUM-{pnum_str}"
        # Truncate sheet name to 31 chars (Excel limit)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 20

        # ── Row 1: section headers ──
        ws.merge_cells("A1:E1")
        _apply_header(ws.cell(1, 1), "1.Gage R&R Data",   bg=CLR_HEADER_BG)
        ws.merge_cells("F1:G1")
        _apply_header(ws.cell(1, 6), "2.Graph Report",    bg=CLR_HEADER_BG)

        # ── Row 2: sub-headers ──
        ws.merge_cells("A2:E2")
        _apply_header(ws.cell(2, 1), "3.Conclusion:",    bg=CLR_SECTION_BG, fg="1F4E79", align="left")

        # ── Metrics rows ──
        pct_grr = grr_res.grr_pct if grr_res else 0.0
        pct_pt  = grr_res.pt_ratio  if grr_res else 0.0
        ndc     = grr_res.ndc     if grr_res else 0

        res_bg, res_fg, res_text = _grr_result_color(pct_grr)

        lsl = spec.get("lsl", 0)
        usl = spec.get("usl", 65555)

        metrics = [
            (None,                         None,                        None,                      None,                        None),
            ("%GR&R=",                      round(pct_grr, 2),           None,                      None,                        "a.%GR&R <=10%, Good;  10%<%GR&R<=30% Acceptable; %GR&R>30% Bad."),
            ("%P/T=",                       round(pct_pt, 2),            None,                      None,                        "b.%P/T<=10% Good;  10%<%P/T<=30% Acceptable; %P/T>30% Bad."),
            ("NDC=",                        ndc,                        None,                      None,                        "c.NDC>10 Good; 5<=NDC<=10 Acceptable, NDC<5 Bad"),
            (None,                          None,                       None,                      None,                        None),
            ("LSL=",                         lsl,                        None,                      None,                        None),
            ("USL=",                         usl,                        None,                      None,                        None),
            ("Tolerance=",                    usl - lsl,                  None,                      None,                        None),
        ]

        for i, row_data in enumerate(metrics, start=3):
            for j, val in enumerate(row_data, start=1):
                if val is None:
                    ws.cell(i, j).border = _thin_border()
                    continue
                is_metric_label = j == 1 and val and "=" in str(val)
                is_metric_val   = j == 2 and i in (4, 5, 6)
                is_note         = j == 5

                if is_note:
                    ws.merge_cells(f"E{i}:G{i}")
                    cell = ws.cell(i, j)
                    cell.value = str(val)
                    cell.font  = Font(name="Calibri", size=8, color="595959")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = _thin_border()
                    ws.row_dimensions[i].height = 20
                elif is_metric_label:
                    cell = ws.cell(i, j)
                    cell.value = str(val)
                    cell.font  = Font(name="Calibri", bold=True, size=10, color="1F4E79")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = _thin_border()
                elif is_metric_val and i == 4:  # %GR&R value
                    cell = ws.cell(i, j)
                    cell.value = float(val) / 100 if isinstance(val, (int, float)) and val > 1 else val
                    cell.number_format = "0.00%"
                    cell.font  = Font(name="Calibri", bold=True, size=11, color=res_fg)
                    cell.fill  = _hdr_fill(res_bg)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif is_metric_val and i == 5:  # %P/T value
                    cell = ws.cell(i, j)
                    cell.value = float(val) / 100 if isinstance(val, (int, float)) and val > 1 else val
                    cell.number_format = "0.00%"
                    cell.font  = Font(name="Calibri", bold=True, size=11)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif is_metric_val and i == 6:  # NDC value
                    cell = ws.cell(i, j)
                    cell.value = val
                    cell.font  = Font(name="Calibri", bold=True, size=11,
                                      color="1F4E79" if ndc >= 10 else "9C6500")
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell = ws.cell(i, j)
                    cell.value = val
                    cell.font  = Font(name="Calibri", size=10)
                    cell.border = _thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        cell.number_format = "#,##0"

        # ── Result row ──
        res_row = 11
        ws.merge_cells(f"A{res_row}:D{res_row}")
        _apply_header(ws.cell(res_row, 1), "GR&R Result:", bg=res_bg, fg=res_fg, align="left")
        ws.merge_cells(f"E{res_row}:G{res_row}")
        cell_result = ws.cell(res_row, 5)
        cell_result.value = res_text
        cell_result.font  = Font(name="Calibri", bold=True, size=11, color=res_fg)
        cell_result.fill  = _hdr_fill(res_bg)
        cell_result.border= _thin_border()
        cell_result.alignment = Alignment(horizontal="center", vertical="center")

        # ── Graph Report: embed image ──
        if chart_img_path and os.path.isfile(chart_img_path):
            try:
                img = XLImage(chart_img_path)
                img.height = 220
                img.width  = 340
                ws.add_image(img, "F2")
                ws.row_dimensions[2].height = 120
                ws.row_dimensions[3].height = 120
            except Exception as eg:
                logger.warning("Could not embed GRR chart: %s", eg)
        else:
            # Placeholder text in graph area
            ws.merge_cells("F2:G10")
            cell_g = ws.cell(2, 6)
            cell_g.value = "(Chart: GageRR graph)"
            cell_g.font  = Font(name="Calibri", size=9, color="AAAAAA", italic=True)
            cell_g.alignment = Alignment(horizontal="center", vertical="center")

        # ── EV / PV / TV detail table ──
        detail_row = 13
        ws.merge_cells(f"A{detail_row}:G{detail_row}")
        _apply_header(ws.cell(detail_row, 1), "GRR Component Analysis", bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG, align="left")

        comp_headers = ["Component", "Value", "Unit", "", "", "", ""]
        for j, h in enumerate(comp_headers, start=1):
            _apply_header(ws.cell(detail_row + 1, j), h, bg=CLR_SECTION_BG, fg="1F4E79", bold=True)

        if grr_res:
            comp_data = [
                ("EV (Repeatability)",    round(grr_res.ev, 2)  if grr_res.ev  else "-", "Lumen"),
                ("PV (Reproducibility)",  round(grr_res.pv, 2)  if grr_res.pv  else "-", "Lumen"),
                ("TV (Total Variation)",  round(grr_res.tv, 2)  if grr_res.tv  else "-", "Lumen"),
                ("GR&R",                  round(grr_res.tv, 2) if grr_res.tv else "-", "Lumen"),
                ("%GR&R",                 round(pct_grr, 2), "%"),
                ("%P/T",                  round(pct_pt, 2),  "%"),
                ("NDC",                   ndc, "-"),
            ]
        else:
            comp_data = []

        for i2, (comp, val, unit) in enumerate(comp_data, start=detail_row + 2):
            bg = CLR_ALT_ROW if i2 % 2 == 0 else "FFFFFF"
            _apply_data(ws.cell(i2, 1), comp, bg=bg, align="left")
            cell_v = ws.cell(i2, 2)
            cell_v.value = val
            cell_v.font  = Font(name="Calibri", bold=True, size=10)
            cell_v.fill  = _hdr_fill(bg)
            cell_v.border= _thin_border()
            cell_v.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(val, float):
                cell_v.number_format = "0.00"
            _apply_data(ws.cell(i2, 3), unit, bg=bg)

        # ── Footer ──
        footer_row = detail_row + 2 + len(comp_data) + 1
        ws.merge_cells(f"A{footer_row}:G{footer_row}")
        _apply_data(ws.cell(footer_row, 1),
                    "This Template Link to #WI-002388",
                    align="left", bold=True, fg="595959")

        ws.freeze_panes = "A2"


# Reverse mapping for LED → PNUM
PNUM_MAPPING_REV = {v: k for k, v in PNUM_MAPPING.items()}
