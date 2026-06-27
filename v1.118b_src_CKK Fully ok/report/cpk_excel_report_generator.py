# ──────────────────────────────────────────────
#  CPK Excel Report Generator
#  Produces M600099 CPK Excel report
# ──────────────────────────────────────────────
import logging
from datetime import datetime
import os
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from utils.config import LED_SPECS

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────
CLR_HEADER_BG   = "1F4E79"
CLR_HEADER_FG   = "FFFFFF"
CLR_SUBHDR_BG    = "2E75B6"
CLR_SUBHDR_FG    = "FFFFFF"
CLR_SECTION_BG   = "D6E4F0"
CLR_ALT_ROW      = "EBF3FB"
CLR_GRN          = "C6EFCE"
CLR_YLW          = "FFEB9C"
CLR_RED          = "FFC7CE"

def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _bold_border() -> Border:
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG,
         size=10, bold=True, wrap=False, align="center"):
    cell.value = text
    cell.font  = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill   = _fill(bg)
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def _dat(cell, value, fmt=None, bold=False, align="center",
         bg=None, fg="000000"):
    cell.value = value
    cell.font  = Font(name="Calibri", bold=bold, color=fg, size=10)
    if bg:
        cell.fill = _fill(bg)
    cell.border     = _thin_border()
    cell.alignment  = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt

def _cpk_color(cpk: float) -> tuple:
    """Return (bg, fg, label) based on CPK value."""
    if cpk >= 1.67:
        return CLR_GRN, "375623", "Excellent"
    elif cpk >= 1.33:
        return CLR_GRN, "000000", "Good"
    elif cpk >= 1.0:
        return CLR_YLW, "7F6000", "Acceptable"
    else:
        return CLR_RED, "9C0006", "Bad"

def _cpk_pass_fail(cpk: float) -> str:
    if cpk >= 1.33:
        return "PASS"
    elif cpk >= 1.0:
        return "ACCEPTABLE"
    return "FAIL"


class CPKExcelReportGenerator:
    """
    Generates a CPK Excel report matching the HVTE-M600099_MT7_CPK template.

    Sheet structure:
      1. M600099 Config A data — raw measurement data
      2. M600099 CPK         — CPK summary with Minitab diagram placeholders
    """

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, results: Dict, df: pd.DataFrame,
                 inline_charts: bool = False,
                 minitab_path: Optional[str] = None,
                 product_name: str = "Beta Charger",
                 station: str = "MT7") -> str:
        """
        Generate CPK Excel report.

        Args:
            results:     {item_name: {"cpk": CPKResult}}
            df:         Full CPK DataFrame (Config A raw data sheet)
            inline_charts: Whether to embed Minitab chart images
            minitab_path:  Path to Minitab Mtb.exe
            product_name:   Product name for report header
            station:        Station ID

        Returns:
            Path to the saved .xlsx file
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── 1. Raw data sheet ──
        self._make_data_sheet(wb, df, product_name)

        # ── 2. CPK summary sheet ──
        self._make_cpk_sheet(wb, results, inline_charts)

        # ── Save ──
        filename = (
            f"CPK_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        out_path = self.output_dir / filename
        wb.save(out_path)
        logger.info(f"CPK Excel report saved: {out_path}")
        return str(out_path)

    # ── Raw Data Sheet ────────────────────────────

    def _make_data_sheet(self, wb, df: pd.DataFrame, product_name: str = "Beta Charger"):
        ws = wb.create_sheet("raw data")

        # Determine dynamic columns FIRST (needed before title merge)
        meta_cols = {"sn", "appraiser", "Config", "TBB_UID", "QR_SCAN", "qr_scan",
                     "sample", "part", "Serial", "SerialNumber"}
        led_cols = [c for c in df.columns
                    if c not in meta_cols
                    and not c.lower().startswith("unnamed")
                    and pd.api.types.is_numeric_dtype(df[c])]
        headers = ["sn", "QR_SCAN", "Config", "TBB_UID"] + sorted(led_cols)
        n_cols = len(headers)

        # Column widths (dynamic based on n_cols)
        ws.column_dimensions["A"].width = 22   # sn
        ws.column_dimensions["B"].width = 22   # QR_SCAN
        ws.column_dimensions["C"].width = 12   # Config
        ws.column_dimensions["D"].width = 18   # TBB_UID
        for j in range(5, n_cols + 1):
            ws.column_dimensions[get_column_letter(j)].width = 16

        # Title row removed per user request
        # Header row is now row 1 (was row 2)
        for j, h in enumerate(headers, start=1):
            _hdr(ws.cell(1, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG,
                 size=9, align="center")
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # Filter out all-None rows from df before writing
        df_data = df.dropna(how='all').reset_index(drop=True)

        for i, (idx, row) in enumerate(df_data.iterrows()):
            r = i + 3
            bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"

            # sn
            sn_val = str(row.get("sn", "")) if "sn" in df.columns else \
                     str(row.get("SN", ""))
            _dat(ws.cell(r, 1), sn_val, bg=bg, align="left")

            # QR_SCAN
            qr = str(row.get("QR_SCAN", "")) if "QR_SCAN" in df.columns else \
                 str(row.get("qr_scan", ""))
            _dat(ws.cell(r, 2), qr, bg=bg, align="left")

            # Config
            cfg = str(row.get("Config", "Config A")) if "Config" in df.columns else "Config A"
            _dat(ws.cell(r, 3), cfg, bg=bg)

            # TBB_UID
            uid = str(row.get("TBB_UID", "")) if "TBB_UID" in df.columns else ""
            _dat(ws.cell(r, 4), uid, bg=bg, align="left")

            # LED values — keep original type from Jason log, convert to number if possible
            for j, col in enumerate(led_cols, start=5):
                val = row.get(col, None)
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    cell_val = None
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell_val = val  # keep original int/float type
                else:
                    # try to parse as number
                    try:
                        cell_val = float(val)
                        # keep as int if it's a whole number
                        if cell_val == int(cell_val):
                            cell_val = int(cell_val)
                    except (ValueError, TypeError):
                        cell_val = str(val)  # keep as string if not a number
                _dat(ws.cell(r, j), cell_val, bg=bg, fmt="General")

        # ── Summary stats at bottom ──
        last_data_row = len(df_data) + 1  # was +3 (title+header rows removed)
        stat_row = last_data_row + 2

        ws.merge_cells(f"A{stat_row}:{get_column_letter(n_cols)}{stat_row}")
        _hdr(ws.cell(stat_row, 1), "Statistical Summary",
             bg=CLR_SECTION_BG, fg="1F4E79", align="left", size=10)

        stat_row += 1
        stat_headers = ["", "", "", ""] + ["Min", "Max", "Mean", "Std"]
        for j, h in enumerate(stat_headers, start=1):
            c = ws.cell(stat_row, j)
            if h:
                c.value = h
                c.font = Font(name="Calibri", bold=True, size=9, color="1F4E79")
                c.fill = _fill(CLR_SECTION_BG)
                c.border = _thin_border()
                c.alignment = Alignment(horizontal="center", vertical="center")

        stat_row += 1
        for j, col in enumerate(led_cols, start=5):
            vals = pd.to_numeric(df_data[col], errors="coerce").dropna()
            bg = CLR_ALT_ROW if j % 2 == 0 else "FFFFFF"
            if len(vals) > 0 and not vals.isna().all():
                mn = vals.min(); mx = vals.max(); av = vals.mean(); sd = vals.std()
                _dat(ws.cell(stat_row, j),     round(mn, 2) if not pd.isna(mn) else None, bg=bg, fmt="0.00")
                _dat(ws.cell(stat_row, j + 1), round(mx, 2) if not pd.isna(mx) else None, bg=bg, fmt="0.00")
                _dat(ws.cell(stat_row, j + 2), round(av, 2) if not pd.isna(av) else None, bg=bg, fmt="0.00")
                _dat(ws.cell(stat_row, j + 3), round(sd, 2) if not pd.isna(sd) else None, bg=bg, fmt="0.00")
            break  # only one pass needed for summary row

    # ── CPK Summary Sheet ────────────────────────

    def _make_cpk_sheet(self, wb, results: Dict, inline_charts: bool = False):
        ws = wb.create_sheet("CPK")

        # Column widths
        ws.column_dimensions["A"].width = 10   # PUNM
        ws.column_dimensions["B"].width = 30   # Test Item
        ws.column_dimensions["C"].width = 14   # Low Limit
        ws.column_dimensions["D"].width = 14   # High Limit
        ws.column_dimensions["E"].width = 12   # CPK
        ws.column_dimensions["F"].width = 30   # Minitab Diagram

        # Title
        ws.merge_cells("A1:F1")
        _hdr(ws.cell(1, 1),
             f"CPK Analysis Report — {datetime.now():%B %d, %Y}",
             bg=CLR_HEADER_BG, size=12, bold=True)
        ws.row_dimensions[1].height = 24

        # Header row
        headers = ["PUNM", "Test Item", "Low Limit", "High Limit", "CPK", "Minitab Diagram"]
        for j, h in enumerate(headers, start=1):
            _hdr(ws.cell(2, j), h, bg=CLR_SUBHDR_BG, fg=CLR_SUBHDR_FG,
                 size=10, wrap=True)
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        # Data rows — derive items dynamically from results keys (actual analyzed items)
        cpk_items = [(None, item) for item in sorted(results.keys())]

        for i, (pnum, item) in enumerate(cpk_items):
            r = i + 3
            cpk_res = results.get(item, {}).get("cpk")
            # Use LSL/USL from cpk_result (parsed from JSON), fall back to LED_SPECS for LEDs
            if cpk_res and (cpk_res.lsl is not None or cpk_res.usl is not None):
                lsl = cpk_res.lsl
                usl = cpk_res.usl
            else:
                spec = LED_SPECS.get(item, {})
                lsl = spec.get("lsl", 0)
                usl = spec.get("usl", 65555)
            cpk_val = round(cpk_res.cpk, 2) if cpk_res and cpk_res.cpk is not None else None

            bg, fg, label = _cpk_color(cpk_val or 0)
            pass_fail = _cpk_pass_fail(cpk_val or 0)

            row_bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"

            _dat(ws.cell(r, 1), pnum,         bg=row_bg, fmt="0", align="center")
            _dat(ws.cell(r, 2), item,          bg=row_bg, align="left")
            _dat(ws.cell(r, 3), lsl, bg=row_bg, fmt="General")
            _dat(ws.cell(r, 4), usl, bg=row_bg, fmt="General")

            # CPK value cell with colour
            cpk_cell = ws.cell(r, 5)
            cpk_cell.value = cpk_val
            cpk_cell.number_format = "0.00"
            cpk_cell.font  = Font(name="Calibri", bold=True, color=fg, size=11)
            cpk_cell.fill  = _fill(bg)
            cpk_cell.border= _thin_border()
            cpk_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Minitab diagram: embed image or show text
            diag_cell = ws.cell(r, 6)
            diag_cell.border = _thin_border()
            if cpk_val is not None:
                chart_path = (cpk_res.chart_paths.get("capability")
                             if cpk_res and hasattr(cpk_res, 'chart_paths') else None)
                if inline_charts and chart_path and os.path.isfile(chart_path):
                    try:
                        img = XLImage(chart_path)
                        # Cap image to fit column F (30 chars ≈ 180 px) and a reasonable row height
                        # WIDE chart typical of Minitab capability plots: cap width at 175 px, scale height
                        img.width  = 175
                        img.height = 115
                        ws.add_image(img, f"F{r}")
                        diag_cell.value = "(Chart embedded)"
                        diag_cell.font  = Font(name="Calibri", size=7, color="595959", italic=True)
                        diag_cell.fill  = _fill(bg)
                        diag_cell.alignment = Alignment(horizontal="center", vertical="center")
                        ws.row_dimensions[r].height = 90   # enough for 115 px image
                    except Exception as eg:
                        diag_cell.value = f"{label}  |  {pass_fail}"
                        diag_cell.font  = Font(name="Calibri", bold=True, color=fg, size=9)
                        diag_cell.fill  = _fill(bg)
                        diag_cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    diag_cell.value = f"{label}  |  {pass_fail}"
                    diag_cell.font  = Font(name="Calibri", bold=True, color=fg, size=9)
                    diag_cell.fill  = _fill(bg)
                    diag_cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                diag_cell.value = "N/A"
                diag_cell.font  = Font(name="Calibri", color="AAAAAA", size=9)
                diag_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Notes ──
        note_row = len(cpk_items) + 5
        ws.merge_cells(f"A{note_row}:F{note_row}")
        note_text = (
            "Notes:  "
            "CPK >= 1.67 → Excellent;  "
            "1.33 <= CPK < 1.67 → Good;  "
            "1.00 <= CPK < 1.33 → Acceptable;  "
            "CPK < 1.00 → Bad.  "
            "Low Limit / High Limit = specification limits from LED spec.  "
            "Minitab Diagram column: embed chart images if Minitab is configured."
        )
        note_cell = ws.cell(note_row, 1)
        note_cell.value = note_text
        note_cell.font  = Font(name="Calibri", italic=True, size=8, color="595959")
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[note_row].height = 35

        ws.freeze_panes = "A3"
