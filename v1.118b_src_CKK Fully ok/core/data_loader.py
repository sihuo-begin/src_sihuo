# ──────────────────────────────────────────────
#  Data Loader
#  Reads GRR / CPK Excel files exported from MT7
# ──────────────────────────────────────────────
import logging
from pathlib import Path
from typing import Dict, List

import pandas as pd
import openpyxl
from utils.config import LED_MAPPING

logger = logging.getLogger(__name__)


class DataLoader:
    """
    Auto-detects two GRR/CPK Excel formats:

    Format A – MT7 raw log (e.g. log_20260505153529):
      - Single sheet, columns include: sn, QR_SCAN, LED_XXX_INTENSITY, status, etc.
      - sn = device serial, QR_SCAN = operator/appraiser

    Format B – FORM-004090 GRR template (e.g. Beta-GRR-Charger):
      - Has a 'GRR From' sheet with columns: Sample, Inspector, PNUM-4024, PNUM-4028, ...
      - 'Summary' sheet with LSL/USL/%GRR/%PT/NDC pre-filled
      - One row per (Sample, Inspector) combination; 3 reps per inspector implied
    """

    # Known PNUM columns in FORM-004090 template
    PNUM_COLS = [
        "PNUM-4024", "PNUM-4028", "PNUM-4032", "PNUM-4036",
        "PNUM-4040", "PNUM-4044", "PNUM-4048", "PNUM-4052", "PNUM-4056"
    ]
    # Map PNUM → LED column name
    PNUM_TO_LED = {
        "PNUM-4024": "LED_RED_D303_INTENSITY",
        "PNUM-4028": "LED_RED_D304_INTENSITY",
        "PNUM-4032": "LED_WHITE_D305_INTENSITY",
        "PNUM-4036": "LED_WHITE_D306_INTENSITY",
        "PNUM-4040": "LED_WHITE_D307_INTENSITY",
        "PNUM-4044": "LED_WHITE_D308_INTENSITY",
        "PNUM-4048": "LED_WHITE_D309_INTENSITY",
        "PNUM-4052": "LED_AMBER_D310_INTENSITY",
        "PNUM-4056": "LED_WHITE_D311_INTENSITY",
    }

    def __init__(self, path: str):
        self.path = Path(path)
        self._load()

    def _load(self):
        """
        Load all sheets.  GRR From has 9 rows of metadata before the
        real column header (row 9 = Sample/Inspector/PNUM-XXXX), so we
        probe with header=9 (0-indexed). All other sheets use header=0.
        """
        wb = openpyxl.load_workbook(self.path, data_only=True)
        self._sheets: Dict[str, pd.DataFrame] = {}

        # Probe GRR From: real header is at row index 8 (row 9 visually)
        grr_header_8 = None
        if "GRR From" in wb.sheetnames:
            probe = pd.read_excel(self.path, sheet_name="GRR From",
                                  header=9, dtype=str)
            probe_cols = [str(c).strip() for c in probe.columns]
            if any(p in probe_cols for p in self.PNUM_COLS):
                grr_header_8 = probe
                logger.info(f"GRR From: PNUM header found at row 9 (0-indexed=8)")

        for name in wb.sheetnames:
            if name == "GRR From" and grr_header_8 is not None:
                df = grr_header_8.copy()
                logger.info(f"Sheet [{name}]: header=9 OK, {len(df)} rows x {len(df.columns)} cols")
            else:
                df = pd.read_excel(self.path, sheet_name=name, dtype=str)
                df.columns = df.columns.str.strip()
                logger.info(f"Sheet [{name}]: {len(df)} rows x {len(df.columns)} cols")
            self._sheets[name] = df

        self.sheet_names = list(self._sheets.keys())
        self.df_raw     = self._sheets
        self._current_sheet = self.sheet_names[0]

        # Auto-detect format first, then set df appropriately
        self.format = self._detect_format()

        # For GRR_TEMPLATE: df = normalized version with LED column names
        # For MT7_LOG:     df = as-is (column names already correct)
        if "GRR From" in self._sheets:
            self.df = self.to_grr_format()
        else:
            self.df = self._sheets[self.sheet_names[0]]
        logger.info(f"Detected format: {self.format}")

    def _detect_format(self) -> str:
        """Detect whether this is Format A (MT7 raw log) or Format B (GRR template)."""
        first_sheet = list(self._sheets.keys())[0]
        cols = self._sheets[first_sheet].columns.tolist()

        # Format B: GRR From sheet exists and has PNUM columns
        if "GRR From" in self._sheets:
            grr_cols = self._sheets["GRR From"].columns.tolist()
            if any(p in grr_cols for p in self.PNUM_COLS):
                return "GRR_TEMPLATE"    # Format B

        # Format A: has LED intensity columns
        if any("INTENSITY" in c for c in cols):
            return "MT7_LOG"              # Format A
        # Default
        return "MT7_LOG"

    @staticmethod
    def _read_sheet(raw: Dict[str, pd.DataFrame], sheet_name: str) -> pd.DataFrame:
        df = raw[sheet_name].copy()
        df.dropna(how="all", inplace=True)
        for col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        return df

    def to_grr_format(self) -> pd.DataFrame:
        """
        Normalize data to a consistent GRR format regardless of source.
        Returns a DataFrame with columns: sn, appraiser, <LED_col>, ...

        Format A: sn=SN, appraiser=QR_SCAN
        Format B: sn=Sample (as str), appraiser=Inspector, LED cols renamed from PNUM
        """
        if self.format == "GRR_TEMPLATE":
            return self._to_grr_template_format()
        else:
            return self._to_mt7_format()

    def _to_grr_template_format(self) -> pd.DataFrame:
        """Convert FORM-004090 GRR From sheet to normalized GRR format."""
        grr_sheet = self._sheets.get("GRR From")
        if grr_sheet is None:
            raise ValueError("No 'GRR From' sheet found in GRR template workbook")

        # Keep only rows that have numeric sample values (skip footer rows)
        df = grr_sheet.copy()

        # Identify columns
        has_sample = "Sample" in df.columns
        has_inspector = "Inspector" in df.columns

        if not (has_sample and has_inspector):
            # Try alternate column names
            col_list = df.columns.tolist()
            logger.warning(f"Expected columns not found. Available: {col_list[:10]}")

        # Build normalized DataFrame
        rows = []
        for _, row in df.iterrows():
            try:
                sample     = str(row.get("Sample", ""))
                inspector  = str(row.get("Inspector", ""))
                # Skip non-data rows
                if not sample.strip() or not inspector.strip():
                    continue
                # Skip rows where Sample column itself contains "Parameter" (header row)
                if sample.lower() in ("", "nan", "sample", "parameter"):
                    continue
                if inspector.lower() in ("", "nan", "inspector", "parameter"):
                    continue
                sample_num = int(float(sample))
                inspector_num = inspector.strip()
            except (ValueError, TypeError):
                continue   # skip footer/metadata rows

            for pnum_col in self.PNUM_COLS:
                if pnum_col not in df.columns:
                    continue
                val = row.get(pnum_col)
                led_col = self.PNUM_TO_LED[pnum_col]
                try:
                    rows.append({
                        "sn":       str(sample_num),
                        "appraiser": inspector_num,
                        led_col:    float(val),
                    })
                except (ValueError, TypeError):
                    pass

        if not rows:
            raise ValueError("No valid GRR data found in 'GRR From' sheet")

        # Keep ALL rows (90 = 10 parts × 3 operators × 3 trials per PNUM).
        # The GRR computation uses n_reps = 3 correctly.
        result = pd.DataFrame(rows)
        # Ensure LED columns are numeric
        for led in self.PNUM_TO_LED.values():
            if led in result.columns:
                result[led] = pd.to_numeric(result[led], errors="coerce")
        return result

    def _to_mt7_format(self) -> pd.DataFrame:
        """
        Normalize MT7 log sheet: prefer 'Sample' for part SN, 'Tester' for operator.
        Skip rows where Sample == 'Parameter' (GRR From header row embedded in MT7).
        """
        sheet_name = getattr(self, "_current_sheet", self.sheet_names[0])
        sheet = self._sheets.get(sheet_name, list(self._sheets.values())[0])
        cols  = sheet.columns.tolist()
        # Resolve sn_col / op_col preferring GRR Template convention
        sn_col = next((c for c in cols if c.lower() == "sample"), None)
        if sn_col is None:
            for c in cols:
                if c.lower() in ("sn", "serial", "device", "lot", "pn"):
                    sn_col = c; break
        op_col = next((c for c in cols if c.lower() == "tester"), None)
        if op_col is None:
            for c in cols:
                if c.lower() in ("appraiser", "inspector", "operator", "qr_scan"):
                    op_col = c; break

        rows = []
        for _, row in sheet.iterrows():
            # Skip rows where Sample == "Parameter" (GRR From header embedded in MT7)
            sn_val = str(row.get(sn_col, "")).strip() if sn_col else ""
            op_val = str(row.get(op_col, "")).strip() if op_col else ""
            if sn_val.lower() in ("", "nan", "sample", "parameter"):
                continue
            if op_val.lower() in ("", "nan", "tester", "inspector", "parameter"):
                continue
            rec = {"sn": sn_val, "appraiser": op_val}
            for c in cols:
                if c not in (sn_col, op_col):
                    rec[c] = row[c]
            rows.append(rec)
        import pandas as pd
        df = pd.DataFrame(rows)
        for col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col], errors="raise")
            except (ValueError, TypeError):
                pass  # keep original string values
        return df

    def get_all_led_columns(self) -> List[str]:
        """
        Return LED measurement column names found in the active/normed sheet.
        Handles all three naming conventions:
          - MT7 raw log:           LED_XXX_INTENSITY
          - GRR_TEMPLATE:         PNUM-XXXX  (normalized to LED name)
          - Intermediate Excel:   LED_XXX    (bare, no suffix)
        """
        df = self.to_grr_format()
        if self.format == "GRR_TEMPLATE":
            return [c for c in df.columns if c in self.PNUM_COLS]
        else:
            # _INTENSITY suffix (MT7 raw log) OR bare LED name (intermediate Excel)
            led_bare = [k for k in LED_MAPPING.keys()]
            return [c for c in df.columns
                    if "INTENSITY" in c
                    or c in led_bare
                    or c in self.PNUM_TO_LED.values()]

    def get_unique_sns(self) -> List[str]:
        """Return sorted list of unique SNs (parts)."""
        df = self.to_grr_format()
        if "sn" in df.columns:
            return sorted(df["sn"].dropna().astype(str).unique())
        return []

    def get_summary_specs(self) -> dict:
        """
        Extract LSL/USL/GRR/%PT/NDC from the 'Summary' sheet (Format B).
        Returns dict: {PNUM: {lsl, usl, pct_grr, pct_pt, ndc, result}}
        """
        summary_sheet = self._sheets.get("Summary") or self._sheets.get("Summary ")
        if summary_sheet is None:
            return {}

        specs = {}
        for pnum in self.PNUM_COLS:
            # Find row where pnum column matches
            for _, row in summary_sheet.iterrows():
                row_vals = [str(v).strip() for v in row.values if str(v).strip()]
                if pnum in row_vals:
                    try:
                        idx = list(summary_sheet.columns).index(pnum)
                        # Navigate across columns to find related fields
                        lsl = summary_sheet.columns["LowLimit"] if "LowLimit" in summary_sheet.columns else None
                        usl = summary_sheet.columns["UpLimit"] if "UpLimit" in summary_sheet.columns else None
                        specs[pnum] = {
                            "lsl": row.get("LowLimit", None),
                            "usl": row.get("UpLimit", None),
                        }
                    except Exception:
                        pass
        return specs

