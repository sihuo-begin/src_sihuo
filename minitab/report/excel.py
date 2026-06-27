import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class ExcelReport:

    def __init__(self, config):
        self.config = config

    def generate(self, output_path, df, cpk_result, grr_result):

        wb = load_workbook(self.config["paths"]["template"])

        self._fill_summary(wb, cpk_result, grr_result)
        self._insert_images(wb)
        self._fill_raw_data(wb, df)

        wb.save(output_path)

    def _fill_summary(self, wb, cpk_result, grr_result):
        ws = wb["Summary"]

        ws["B8"] = round(cpk_result["cpk"], 3)
        ws["B9"] = round(cpk_result["mean"], 3)
        ws["B10"] = round(cpk_result["std"], 3)

        ws["B13"] = grr_result["grr"]

    def _insert_images(self, wb):
        image_dir = self.config["paths"]["images"]

        try:
            wb["CPK_Chart"].add_image(Image(f"{image_dir}/cpk.jpg"), "A1")
        except:
            print("CPK图缺失")

        try:
            wb["GRR_Chart"].add_image(Image(f"{image_dir}/grr.jpg"), "A1")
        except:
            print("GRR图缺失")

    def _fill_raw_data(self, wb, df):
        ws = wb["Raw_Data"]

        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col)

        for r in range(len(df)):
            for c in range(len(df.columns)):
                ws.cell(row=r+2, column=c+1, value=df.iloc[r, c])

    def insert_cpk_images(wb, cpk_results):

        ws = wb["CPK_Chart"]

        row = 1

        for col, result in cpk_results.items():
            try:
                img = Image(result["image"])
                ws.add_image(img, f"A{row}")
                row += 30
            except:
                print(f"{col} 图缺失")

    def insert_grr_image(wb, grr_result):

        try:
            img = Image(grr_result["image"])
            wb["GRR_Chart"].add_image(img, "A1")
        except:
            print("GRR图缺失")
